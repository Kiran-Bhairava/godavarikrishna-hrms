"""
Attendance System — FastAPI backend with Clean Role-Based Architecture

Architecture:
  Admin (one-time setup)
    ↓ (creates via /api/admin/register-first-admin or /api/admin/register-hr)
  HR Users
    ↓ (creates via /api/hr/employees)
  Employees (full profiles)
    ↓ (login via /api/auth/login)
  Use System
"""
import logging
import math
import pytz
import dotenv
from contextlib import asynccontextmanager
from datetime import datetime, timedelta, date, time
from io import BytesIO
from pathlib import Path
from typing import Optional, Annotated
import asyncpg
from fastapi import FastAPI, Depends, HTTPException, Query, Path as PathParam, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from jose import JWTError, jwt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from passlib.context import CryptContext
from pydantic import BaseModel, EmailStr, field_validator
import re
import asyncio
# ── Local imports ──────────────────────────────────────────────
from config import settings
from db import get_db, init_db, close_db
from auth import get_current_user, require_hr, require_admin, security
from schemas import LoginResponse
from api_credentials import generate_temp_password
from email_utils import send_welcome_credentials

# ── mandate Include routers ──────────────────────────────────
from routers import regularization_router, leave_router, payroll_router,sandwich_router

dotenv.load_dotenv()

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# ── Logging ────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(name)s  %(message)s",
)
logger = logging.getLogger("attendance")


# ══════════════════════════════════════════════════════════════
# APP SETUP
# ══════════════════════════════════════════════════════════════

@asynccontextmanager
async def lifespan(app: FastAPI):
    await init_db()
    logger.info("✓ Database pool initialized")

    scheduler_task = asyncio.create_task(_run_auto_punchout())
    logger.info("✓ Auto punch-out scheduler started")

    yield

    scheduler_task.cancel()
    try:
        await scheduler_task
    except asyncio.CancelledError:
        pass

    await close_db()
    logger.info("✓ Database pool closed")


app = FastAPI(
    title="Attendance System",
    description="SDPL Attendance Manager",
    version="3.0.0",
    lifespan=lifespan,
    redoc_url="/api/redoc",
)

# ── CORS Middleware ────────────────────────────────────────────
if settings.allowed_origins:
    app.add_middleware(
        CORSMiddleware,
        allow_origins=settings.allowed_origins,
        allow_credentials=True,
        allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE"],
        allow_headers=["Authorization", "Content-Type"],
    )

app.include_router(regularization_router)
app.include_router(leave_router)
app.include_router(payroll_router)
app.include_router(sandwich_router)
# ══════════════════════════════════════════════════════════════
# SCHEMAS (LOCAL ONLY)
# ══════════════════════════════════════════════════════════════

class LoginRequest(BaseModel):
    email: EmailStr
    password: str

    @field_validator("password")
    @classmethod
    def not_empty(cls, v):
        if not v.strip():
            raise ValueError("Password must not be empty")
        return v


class RegisterRequest(BaseModel):
    """For admin and HR registration"""
    full_name: str
    email: EmailStr
    password: str

    @field_validator("password")
    @classmethod
    def password_length(cls, v):
        if len(v) < 8:
            raise ValueError("Password must be at least 8 characters")
        return v


class PunchRequest(BaseModel):
    latitude: float
    longitude: float

    @field_validator("latitude")
    @classmethod
    def valid_lat(cls, v):
        if not (-90 <= v <= 90):
            raise ValueError("Invalid latitude")
        return v

    @field_validator("longitude")
    @classmethod
    def valid_lng(cls, v):
        if not (-180 <= v <= 180):
            raise ValueError("Invalid longitude")
        return v


class OnboardRequest(BaseModel):
    """For HR onboarding employees"""
    full_name: str
    work_email: EmailStr
    password: Optional[str] = None
    personal_email: Optional[str] = None
    phone: Optional[str] = None
    dob: Optional[str] = None
    gender: Optional[str] = None
    blood_group: Optional[str] = None
    nationality: Optional[str] = None
    home_address: Optional[str] = None
    emg_name: Optional[str] = None
    emg_phone: Optional[str] = None
    emg_rel: Optional[str] = None
    
    job_title: Optional[str] = None
    designation: Optional[str] = None
    department: Optional[str] = None
    sub_department: Optional[str] = None
    grade: Optional[str] = None
    date_of_joining: Optional[str] = None
    branch_id: Optional[int] = None
    l1_manager_id: Optional[int] = None
    l2_manager_id: Optional[int] = None
    role: str = "employee"
    cost_centre: Optional[str] = None
    
    employment_type: Optional[str] = None
    contract_end: Optional[str] = None
    probation_end: Optional[str] = None
    notice_period: Optional[str] = None
    
    shift_start: time = time(9, 0)
    shift_end: time = time(18, 0)
    work_mode: Optional[str] = "On-Site"
    weekly_off: Optional[str] = "Saturday & Sunday"
    work_location: Optional[str] = None
    asset_id: Optional[str] = None
    
    annual_ctc: Optional[float] = None
    pay_frequency: Optional[str] = "Monthly"
    pf_enrolled: bool = True
    esic_applicable: bool = True
    bank_name: Optional[str] = None
    bank_account: Optional[str] = None
    bank_ifsc: Optional[str] = None
    pan_number: Optional[str] = None

    @field_validator("role")
    @classmethod
    def valid_role(cls, v):
        if v not in ("employee", "hr", "admin"):
            raise ValueError("role must be employee, hr, or admin")
        return v

    @field_validator("password")
    @classmethod
    def min_length(cls, v):
        if v is None or v == "":
            return None
        if len(v) < 8:
            raise ValueError("Password must be at least 8 characters")
        return v


class UpdateOnboardingStatus(BaseModel):
    status: str

    @field_validator("status")
    @classmethod
    def valid_status(cls, v):
        if v not in ("awaiting", "in-progress", "completed"):
            raise ValueError("Invalid status")
        return v


class UpdateEmployeeRequest(BaseModel):
    """Partial update — all fields optional."""
    full_name:       Optional[str] = None
    personal_email:  Optional[str] = None
    phone:           Optional[str] = None
    dob:             Optional[str] = None
    gender:          Optional[str] = None
    blood_group:     Optional[str] = None
    nationality:     Optional[str] = None
    home_address:    Optional[str] = None
    emg_name:        Optional[str] = None
    emg_phone:       Optional[str] = None
    emg_rel:         Optional[str] = None
    branch_id:       Optional[int] = None
    job_title:       Optional[str] = None
    designation:     Optional[str] = None
    department:      Optional[str] = None
    sub_department:  Optional[str] = None
    grade:           Optional[str] = None
    date_of_joining: Optional[str] = None
    cost_centre:     Optional[str] = None
    l1_manager_id:   Optional[int] = None
    l2_manager_id:   Optional[int] = None
    employment_type: Optional[str] = None
    contract_end:    Optional[str] = None
    probation_end:   Optional[str] = None
    notice_period:   Optional[str] = None
    shift_start:     Optional[str] = None
    shift_end:       Optional[str] = None
    work_mode:       Optional[str] = None
    weekly_off:      Optional[str] = None
    work_location:   Optional[str] = None
    asset_id:        Optional[str] = None
    annual_ctc:      Optional[float] = None
    pay_frequency:   Optional[str] = None
    pf_enrolled:     Optional[bool] = None
    esic_applicable: Optional[bool] = None
    bank_name:       Optional[str] = None
    bank_account:    Optional[str] = None
    bank_ifsc:       Optional[str] = None
    pan_number:      Optional[str] = None
    role:            Optional[str] = None

    @field_validator("role")
    @classmethod
    def valid_role(cls, v):
        if v is not None and v not in ("employee", "hr", "admin"):
            raise ValueError("role must be employee, hr, or admin")
        return v


# ══════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════

def haversine(lat1, lon1, lat2, lon2) -> float:
    """Calculate distance in meters between two coordinates."""
    R = 6_371_000  # Earth radius in meters
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))


def hash_password(p: str) -> str:
    """Hash a password."""
    return pwd_context.hash(p)


def verify_password(plain: str, hashed: str) -> bool:
    """Verify a password against its hash."""
    return pwd_context.verify(plain, hashed)


def create_token(user_id: int, email: str, role: str) -> str:
    """Create JWT token."""
    exp = datetime.now(tz=pytz.utc) + timedelta(hours=settings.access_token_expire_hours)
    return jwt.encode(
        {"sub": str(user_id), "email": email, "role": role, "exp": exp},
        settings.secret_key, algorithm=settings.algorithm,
    )


def local_now() -> datetime:
    """Get current time in office timezone."""
    return datetime.now(pytz.timezone(settings.office_timezone))


def to_local(dt: Optional[datetime]) -> Optional[datetime]:
    """Convert UTC datetime to local timezone."""
    if dt is None:
        return None
    tz = pytz.timezone(settings.office_timezone)
    if dt.tzinfo is None:
        dt = pytz.utc.localize(dt)
    return dt.astimezone(tz)


def parse_date(s: Optional[str]) -> Optional[date]:
    """Parse date string (YYYY-MM-DD)."""
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def parse_date_param(s: Optional[str]) -> date:
    """Parse date from query param, default to today."""
    if not s:
        return date.today()
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(400, "Expected date format YYYY-MM-DD")


# ── Validation Regexes ─────────────────────────────────────────
PAN_REGEX   = re.compile(r"^[A-Z]{5}[0-9]{4}[A-Z]$")
IFSC_REGEX  = re.compile(r"^[A-Z]{4}0[A-Z0-9]{6}$")
PHONE_REGEX = re.compile(r"^[6-9]\d{9}$")
EMAIL_REGEX = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _validate_common_fields(
    *,
    full_name=None,
    work_email=None,
    personal_email=None,
    phone=None,
    emg_phone=None,
    pan_number=None,
    bank_ifsc=None,
    bank_account=None,
    annual_ctc=None,
    role=None,
    date_of_joining=None,
    probation_end=None,
    contract_end=None,
    require_role=False,
):
    """Validate common employee fields. Single source of truth — do not duplicate."""

    # ── Name ─────────────────────────────
    if full_name is not None and len(full_name.strip()) < 2:
        raise HTTPException(400, "Full name must be at least 2 characters")

    # ── Emails ───────────────────────────
    if work_email is not None and not EMAIL_REGEX.match(work_email):
        raise HTTPException(400, "Invalid work email format")

    if personal_email is not None and not EMAIL_REGEX.match(personal_email):
        raise HTTPException(400, "Invalid personal email format")

    # ── Phones ───────────────────────────
    if phone is not None and not PHONE_REGEX.match(phone):
        raise HTTPException(400, "Phone must be 10 digits starting with 6-9")

    if emg_phone is not None and not PHONE_REGEX.match(emg_phone):
        raise HTTPException(400, "Emergency phone must be valid 10-digit number")

    # ── PAN ──────────────────────────────
    if pan_number is not None:
        pan = pan_number.upper()
        if not PAN_REGEX.match(pan):
            raise HTTPException(400, "Invalid PAN format (ABCDE1234F)")

    # ── IFSC ─────────────────────────────
    if bank_ifsc is not None:
        if not IFSC_REGEX.match(bank_ifsc.upper()):
            raise HTTPException(400, "Invalid IFSC format (e.g., SBIN0001234)")

    # ── Bank account ─────────────────────
    if bank_account is not None and len(bank_account) < 8:
        raise HTTPException(400, "Bank account number seems too short")

    # ── Salary ───────────────────────────
    if annual_ctc is not None and annual_ctc <= 0:
        raise HTTPException(400, "Annual CTC must be greater than 0")

    # ── Role validation ──────────────────
    if role is not None:
        if role not in ("employee", "hr", "admin"):
            raise HTTPException(400, "Role must be employee, hr, or admin")

    if require_role and role not in ("employee", "hr"):
        raise HTTPException(400, "Role must be either 'employee' or 'hr'")

    # ── Date logic (convert safely) ──────
    doj = parse_date(date_of_joining) if date_of_joining else None
    prob = parse_date(probation_end) if probation_end else None
    contract = parse_date(contract_end) if contract_end else None

    if doj and prob and prob < doj:
        raise HTTPException(400, "Probation end cannot be before joining date")

    if doj and contract and contract < doj:
        raise HTTPException(400, "Contract end cannot be before joining date")


def validate_onboard_payload(payload: OnboardRequest):
    """Validate onboard request."""
    _validate_common_fields(
        full_name=payload.full_name,
        work_email=payload.work_email,
        personal_email=payload.personal_email,
        phone=payload.phone,
        emg_phone=payload.emg_phone,
        pan_number=payload.pan_number,
        bank_ifsc=payload.bank_ifsc,
        bank_account=payload.bank_account,
        annual_ctc=payload.annual_ctc,
        role=payload.role,
        date_of_joining=payload.date_of_joining,
        probation_end=payload.probation_end,
        contract_end=payload.contract_end,
        require_role=True,
    )


def validate_update_payload(req: UpdateEmployeeRequest):
    """Validate employee update request."""
    _validate_common_fields(
        full_name=req.full_name,
        personal_email=req.personal_email,
        phone=req.phone,
        emg_phone=req.emg_phone,
        pan_number=req.pan_number,
        bank_ifsc=req.bank_ifsc,
        bank_account=req.bank_account,
        annual_ctc=req.annual_ctc,
        role=req.role,
        date_of_joining=req.date_of_joining,
        probation_end=req.probation_end,
        contract_end=req.contract_end,
    )
# ══════════════════════════════════════════════════════════════
# AUTO PUNCH-OUT SCHEDULER
# ══════════════════════════════════════════════════════════════

async def _run_auto_punchout():
    """
    Daily job: auto punch-out any employee still clocked in at 20:00 local time.
    Runs in a background loop, fires once per day.

    Flags:
      - attendance_logs: is_valid=FALSE  → HR can identify auto-generated punches
      - daily_summary:   payroll_notes   → visible in HR daily report
    """
    import db as _db

    tz = pytz.timezone(settings.office_timezone)

    while True:
        try:
            now    = datetime.now(tz)
            target = now.replace(hour=20, minute=0, second=0, microsecond=0)
            if now >= target:
                target += timedelta(days=1)

            wait_seconds = (target - now).total_seconds()
            logger.info(
                "Auto punch-out: next run in %.0fs (at %s)",
                wait_seconds, target.strftime("%Y-%m-%d %H:%M"),
            )
            await asyncio.sleep(wait_seconds)

            now   = datetime.now(tz)
            today = now.date()

            # Fixed auto punch-out time: 20:00 local, stored as UTC (naive) for DB consistency
            auto_out_local = datetime.combine(today, time(20, 0)).replace(tzinfo=tz)
            auto_out_utc   = auto_out_local.astimezone(pytz.utc).replace(tzinfo=None)

            async with _db.db_pool.acquire() as conn:
                # All employees with an open punch-in today (punch-in exists, no punch-out)
                open_rows = await conn.fetch(
                    """
                    SELECT DISTINCT ON (al.user_id)
                        al.user_id,
                        al.branch_id,
                        ds.first_punch_in
                    FROM attendance_logs al
                    JOIN employees e ON e.user_id = al.user_id
                    LEFT JOIN daily_summary ds
                           ON ds.user_id   = al.user_id
                          AND ds.work_date = $1
                    WHERE (al.punched_at AT TIME ZONE $2)::date = $1
                      AND al.punch_type = 'in'
                      AND e.is_active = TRUE
                      AND NOT EXISTS (
                          SELECT 1 FROM attendance_logs al2
                          WHERE al2.user_id = al.user_id
                            AND (al2.punched_at AT TIME ZONE $2)::date = $1
                            AND al2.punch_type = 'out'
                      )
                    ORDER BY al.user_id, al.punched_at DESC
                    """,
                    today, settings.office_timezone,
                )

                if not open_rows:
                    logger.info("Auto punch-out: no open punch-ins for %s", today)
                    continue

                logger.info(
                    "Auto punch-out: closing %d open punch-in(s) for %s",
                    len(open_rows), today,
                )

                for row in open_rows:
                    user_id   = row["user_id"]
                    branch_id = row["branch_id"]

                    # Minutes = 8 PM minus first_punch_in (0 if punch-in row missing)
                    total_min = 0
                    if row["first_punch_in"]:
                        raw_in = row["first_punch_in"]
                        # first_punch_in is stored timezone-aware (UTC) in the DB
                        if raw_in.tzinfo is not None:
                            raw_in = raw_in.replace(tzinfo=None)
                        total_min = max(0, int(
                            (auto_out_utc - raw_in).total_seconds() / 60
                        ))

                    async with conn.transaction():
                        await conn.execute(
                            """
                            INSERT INTO attendance_logs
                                (user_id, branch_id, punch_type,
                                 latitude, longitude, distance_meters,
                                 is_valid, punched_at)
                            VALUES ($1, $2, 'out', 0, 0, 0, FALSE, $3)
                            """,
                            user_id, branch_id, auto_out_utc,
                        )

                        await conn.execute(
                            """
                            UPDATE daily_summary
                            SET last_punch_out  = $2,
                                total_minutes   = CASE WHEN is_regularized
                                                    THEN total_minutes
                                                    ELSE $3 END,
                                payroll_minutes = CASE WHEN is_regularized
                                                    THEN payroll_minutes
                                                    ELSE $3 END,
                                status          = 'present',
                                payroll_status  = CASE WHEN is_regularized
                                                    THEN payroll_status
                                                    ELSE 'present' END,
                                payroll_notes   = CASE WHEN is_regularized
                                                    THEN payroll_notes
                                                    ELSE 'Auto punch-out at 20:00 — verify if needed'
                                                END
                            WHERE user_id = $1 AND work_date = $4
                            """,
                            user_id, auto_out_utc, total_min, today,
                        )

                    logger.info(
                        "Auto punch-out done: user_id=%s total_min=%d",
                        user_id, total_min,
                    )

        except asyncio.CancelledError:
            logger.info("Auto punch-out scheduler stopped")
            break
        except Exception as e:
            logger.error("Auto punch-out error: %s", e, exc_info=True)
            await asyncio.sleep(3600)  # back off 1h on unexpected error, retry

# ══════════════════════════════════════════════════════════════
# AUTH ENDPOINTS
# ══════════════════════════════════════════════════════════════

@app.post("/api/auth/register-first-admin")
async def register_first_admin(
    payload: RegisterRequest,
    db: asyncpg.Connection = Depends(get_db),
):
    """
    One-time endpoint to create the first admin user.
    ⚠️ Only works if NO admin users exist in the system.
    
    After first admin is created, this endpoint is disabled.
    """
    
    # Check if any admin already exists
    admin_exists = await db.fetchrow(
        "SELECT id FROM users WHERE role='admin' LIMIT 1"
    )
    
    if admin_exists:
        raise HTTPException(403, "Admin user already exists. Use /api/admin/register-hr to create HR users.")
    
    # Check for duplicate email
    existing = await db.fetchrow(
        "SELECT id FROM users WHERE LOWER(email) = LOWER($1)",
        payload.email,
    )
    if existing:
        raise HTTPException(409, "User with this email already exists")
    
    # Hash password
    hashed = pwd_context.hash(payload.password)
    
    # Create first admin user
    user = await db.fetchrow(
        """
        INSERT INTO users (email, password_hash, full_name, role, is_active, created_at)
        VALUES ($1, $2, $3, 'admin', TRUE, NOW())
        RETURNING id, email, full_name, role, created_at
        """,
        payload.email,
        hashed,
        payload.full_name,
    )
    
    logger.info("First admin user created: user_id=%s email=%s", user["id"], user["email"])
    
    return {
        "user_id": user["id"],
        "email": user["email"],
        "full_name": user["full_name"],
        "role": user["role"],
        "message": "First admin user created successfully",
        "created_at": user["created_at"].isoformat(),
    }


@app.post("/api/admin/register-hr")
async def register_hr(
    payload: RegisterRequest,
    admin: dict = Depends(require_admin),
    db: asyncpg.Connection = Depends(get_db),
):
    """
    Admin-only endpoint to register new HR users.
    
    Requires: Admin authentication token
    """
    
    # Check for duplicate email
    existing = await db.fetchrow(
        "SELECT id FROM users WHERE LOWER(email) = LOWER($1)",
        payload.email,
    )
    if existing:
        raise HTTPException(409, "User with this email already exists")
    
    # Hash password
    hashed = pwd_context.hash(payload.password)
    
    # Create HR user
    user = await db.fetchrow(
        """
        INSERT INTO users (email, password_hash, full_name, role, is_active, created_at)
        VALUES ($1, $2, $3, 'hr', TRUE, NOW())
        RETURNING id, email, full_name, role, created_at
        """,
        payload.email,
        hashed,
        payload.full_name,
    )
    
    logger.info(
        "HR user created: user_id=%s email=%s by admin_id=%s",
        user["id"], user["email"], admin["id"]
    )
    
    return {
        "user_id": user["id"],
        "email": user["email"],
        "full_name": user["full_name"],
        "role": user["role"],
        "message": "HR user created successfully",
        "created_at": user["created_at"].isoformat(),
    }


@app.post("/api/auth/login", response_model=LoginResponse)
async def login(
    payload: LoginRequest,
    db: asyncpg.Connection = Depends(get_db),
):
    """
    Login endpoint for all users (admin, HR, employees)
    """
    # 1. Look up user via personal_email (employees table) — primary login identifier
    user = await db.fetchrow(
        """
        SELECT u.id, u.email, u.password_hash, u.full_name, u.role,
               u.is_active, u.must_reset_password
        FROM employees e
        JOIN users u ON u.id = e.user_id
        WHERE LOWER(e.personal_email) = LOWER($1)
          AND e.is_active = TRUE
        """,
        payload.email,
    )

    # HR/admin may not have an employee row — fall back to work email
    if not user:
        user = await db.fetchrow(
            """
            SELECT id, email, password_hash, full_name, role, is_active, must_reset_password
            FROM users
            WHERE LOWER(email) = LOWER($1)
              AND role IN ('hr', 'admin')
            """,
            payload.email,
        )

    if not user:
        raise HTTPException(401, "Invalid email or password")

    if not user["is_active"]:
        raise HTTPException(403, "Account is inactive. Contact HR.")

    # 2. Verify password
    if not pwd_context.verify(payload.password, user["password_hash"]):
        raise HTTPException(401, "Invalid email or password")

    user_id = user["id"]

    # 3. Get employee profile (optional for non-employees)
    employee = None
    if user["role"] in ("employee", "hr"):
        employee = await db.fetchrow(
            """
            SELECT e.id, e.branch_id, e.onboarding_status, e.shift_start, e.shift_end,
                   b.name AS branch_name,
                   b.city AS branch_city,
                   b.latitude, b.longitude, b.radius_meters
            FROM employees e
            LEFT JOIN branches b ON b.id = e.branch_id
            WHERE e.user_id = $1
            """,
            user_id,
        )
        
        # Check onboarding status if employee exists
        if employee and employee["onboarding_status"] != "completed":
            raise HTTPException(
                403,
                f"Onboarding incomplete. Status: {employee['onboarding_status']}. Contact HR."
            )

    # 4. Create JWT token — embed must_reset so the frontend and API can both enforce it
    expire = datetime.utcnow() + timedelta(hours=settings.access_token_expire_hours)
    token_data = {
        "sub": str(user_id),
        "role": user["role"],
        "must_reset": bool(user["must_reset_password"]),
        "exp": expire,
    }
    token = jwt.encode(token_data, settings.secret_key, algorithm=settings.algorithm)

    # 5. Update last login
    await db.execute(
        "UPDATE users SET last_login=NOW() WHERE id=$1",
        user_id,
    )

    # 6. Build response
    shift_start = None
    shift_end = None
    if employee and employee["shift_start"]:
        shift_start = employee["shift_start"].strftime("%H:%M")
    if employee and employee["shift_end"]:
        shift_end = employee["shift_end"].strftime("%H:%M")

    user_public = {
        "id": user_id,
        "email": user["email"],
        "full_name": user["full_name"],
        "role": user["role"],
        "branch_id": employee["branch_id"] if employee else None,
        "shift_start": shift_start,
        "shift_end": shift_end,
        "branch_name": employee["branch_name"] if employee else None,
        "branch_city": employee["branch_city"] if employee else None,
        "branch_lat": float(employee["latitude"]) if employee and employee["latitude"] else None,
        "branch_lng": float(employee["longitude"]) if employee and employee["longitude"] else None,
        "radius_meters": employee["radius_meters"] if employee else None,
    }

    return {
        "access_token": token,
        "token_type": "bearer",
        "user": user_public,
        "must_reset_password": user["must_reset_password"] or False,
    }


@app.post("/api/auth/logout")
async def logout():
    """Logout endpoint (token invalidation handled on client)"""
    return {"message": "Logged out"}


class ChangePasswordRequest(BaseModel):
    new_password: str

    @field_validator("new_password")
    @classmethod
    def min_length(cls, v):
        if len(v) < 8:
            raise ValueError("Password must be at least 8 characters")
        return v


@app.post("/api/auth/change-password")
async def change_password(
    payload: ChangePasswordRequest,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: asyncpg.Connection = Depends(get_db),
):
    """
    Change password — works even when must_reset_password=True.
    Does NOT go through get_current_user (which blocks must_reset tokens).
    User is identified from the JWT token — no old password needed on first reset.
    Returns a fresh token (without must_reset) so the frontend can continue without re-login.
    """
    token = credentials.credentials
    try:
        token_payload = jwt.decode(token, settings.secret_key, algorithms=[settings.algorithm])
        user_id = int(token_payload.get("sub"))
        email   = token_payload.get("email")
        role    = token_payload.get("role")
    except (JWTError, ValueError):
        raise HTTPException(401, "Invalid token")

    new_hash = pwd_context.hash(payload.new_password)
    async with db.transaction():
        result = await db.execute(
            "UPDATE users SET password_hash=$1, must_reset_password=FALSE WHERE id=$2",
            new_hash, user_id,
        )
        if result == "UPDATE 0":
            raise HTTPException(404, "User not found")
        await db.execute(
            """INSERT INTO credential_audits (user_id, action, is_temporary, created_at)
               VALUES ($1, 'changed', FALSE, NOW())""",
            user_id,
        )

    # Issue a clean token (must_reset removed) so frontend can navigate without re-login
    fresh_token = create_token(user_id, email, role)

    logger.info("Password changed: user_id=%s", user_id)
    return {
        "message": "Password changed successfully.",
        "access_token": fresh_token,
        "role": role,
    }


@app.get("/api/auth/me")
async def get_me(
    user: dict = Depends(get_current_user),
    db: asyncpg.Connection = Depends(get_db),
):
    """Get current user profile."""
    return user
# ══════════════════════════════════════════════════════════════
# ATTENDANCE
# ══════════════════════════════════════════════════════════════

async def _today_punches(db, user_id: int) -> list[str]:
    rows = await db.fetch(
        """SELECT punch_type FROM attendance_logs
           WHERE user_id = $1
             AND (punched_at AT TIME ZONE $2)::date = (NOW() AT TIME ZONE $2)::date
           ORDER BY punched_at""",
        user_id, settings.office_timezone,
    )
    return [r["punch_type"] for r in rows]


@app.post("/api/attendance/punch-in")
async def punch_in(
    req: PunchRequest,
    user: dict = Depends(get_current_user),
    db: asyncpg.Connection = Depends(get_db),
):
    if not user["branch_id"]:
        raise HTTPException(400, "No branch assigned. Contact HR.")

    # Validate branch coordinates first
    if user["branch_lat"] is None or user["branch_lng"] is None:
        raise HTTPException(400, "Branch location not configured. Contact HR.")

    distance = haversine(
        req.latitude, req.longitude,
        float(user["branch_lat"]), float(user["branch_lng"]),
    )

    radius = user["radius_meters"] or 200
    if distance > radius:
        raise HTTPException(403, f"You are {int(distance)}m away. Must be within {radius}m.")

    punches = await _today_punches(db, user["id"])
    if punches and punches[-1] == "in":
        raise HTTPException(409, "Already punched in.")
    if "in" in punches and "out" in punches:
        raise HTTPException(409, "Attendance already completed for today.")

    now = local_now()
    shift = user["shift_start"]
    is_late, late_min = False, 0
    if shift:
        delta = (datetime.combine(date.today(), now.time())
                 - datetime.combine(date.today(), shift)).total_seconds()
        grace = settings.late_grace_minutes * 60
        if delta > grace:
            is_late, late_min = True, max(0, int((delta - grace) / 60))

    async with db.transaction():
        log = await db.fetchrow(
            """INSERT INTO attendance_logs
               (user_id, branch_id, punch_type, latitude, longitude, distance_meters, is_valid)
               VALUES ($1,$2,'in',$3,$4,$5,TRUE) RETURNING punched_at""",
            user["id"], user["branch_id"], req.latitude, req.longitude, int(distance),
        )
        await db.execute(
            """INSERT INTO daily_summary
            (user_id, work_date, first_punch_in, is_late, late_by_minutes, status, payroll_status)
            VALUES ($1,(NOW() AT TIME ZONE $2)::date,$3,$4,$5,'present','present')
            ON CONFLICT (user_id, work_date) DO UPDATE
            SET first_punch_in=EXCLUDED.first_punch_in,
                is_late=EXCLUDED.is_late,
                late_by_minutes=EXCLUDED.late_by_minutes,
                status='present',
                payroll_status='present'""",
            user["id"], settings.office_timezone, log["punched_at"], is_late, late_min,
        )

    punch_time = to_local(log["punched_at"]).strftime("%I:%M %p")
    return {
        "success": True,
        "message": f"Punched in{f' — late by {late_min} min' if is_late else ''}",
        "time": punch_time,
        "distance": int(distance),
        "is_late": is_late,
        "late_by_minutes": late_min,
        "punched_at_iso": log["punched_at"].isoformat(),
    }


@app.post("/api/attendance/punch-out")
async def punch_out(
    req: PunchRequest,
    user: dict = Depends(get_current_user),
    db: asyncpg.Connection = Depends(get_db),
):
    if not user["branch_id"]:
        raise HTTPException(400, "No branch assigned. Contact HR.")

    if user["branch_lat"] is None or user["branch_lng"] is None:
        raise HTTPException(400, "Branch location not configured. Contact HR.")

    distance = haversine(
        req.latitude, req.longitude,
        float(user["branch_lat"]), float(user["branch_lng"]),
    )

    radius = user["radius_meters"] or 200
    if distance > radius:
        raise HTTPException(403, f"You are {int(distance)}m away. Must be within {radius}m.")

    punches = await _today_punches(db, user["id"])
    if not punches or punches[-1] != "in":
        raise HTTPException(409, "Must punch in first.")

    async with db.transaction():
        log = await db.fetchrow(
            """INSERT INTO attendance_logs
               (user_id, branch_id, punch_type, latitude, longitude, distance_meters, is_valid)
               VALUES ($1,$2,'out',$3,$4,$5,TRUE) RETURNING punched_at""",
            user["id"], user["branch_id"], req.latitude, req.longitude, int(distance),
        )
        summary = await db.fetchrow(
            """SELECT first_punch_in FROM daily_summary
               WHERE user_id=$1 AND work_date=(NOW() AT TIME ZONE $2)::date""",
            user["id"], settings.office_timezone,
        )
        total_min = 0
        if summary and summary["first_punch_in"]:
            total_min = max(0, int((log["punched_at"] - summary["first_punch_in"]).total_seconds() / 60))
        await db.execute(
            """UPDATE daily_summary
            SET last_punch_out  = $2,
                -- Both total_minutes and payroll_minutes are protected on regularized days.
                -- total_minutes is what the calendar view uses to show the gap indicator —
                -- it must stay at the credited value (actual + approved gap), not the raw
                -- punch duration, otherwise the calendar shows a false shortfall.
                total_minutes   = CASE WHEN is_regularized THEN total_minutes   ELSE $3 END,
                payroll_minutes = CASE WHEN is_regularized THEN payroll_minutes ELSE $3 END
            WHERE user_id=$1 AND work_date=(NOW() AT TIME ZONE $4)::date""",
            user["id"], log["punched_at"], total_min, settings.office_timezone,
        )

    punch_time = to_local(log["punched_at"]).strftime("%I:%M %p")
    return {
        "success": True,
        "message": "Punched out. Have a great day!",
        "time": punch_time,
        "total_hours": f"{total_min // 60}h {total_min % 60}m",
    }


@app.get("/api/attendance/status")
async def attendance_status(
    user: dict = Depends(get_current_user),
    db: asyncpg.Connection = Depends(get_db),
):
    logs = await db.fetch(
        """SELECT punch_type, punched_at FROM attendance_logs
           WHERE user_id=$1
             AND (punched_at AT TIME ZONE $2)::date=(NOW() AT TIME ZONE $2)::date
           ORDER BY punched_at""",
        user["id"], settings.office_timezone,
    )
    summary = await db.fetchrow(
        """SELECT * FROM daily_summary
           WHERE user_id=$1 AND work_date=(NOW() AT TIME ZONE $2)::date""",
        user["id"], settings.office_timezone,
    )

    punches = [r["punch_type"] for r in logs]
    last    = logs[-1] if logs else None
    state   = "none"
    if punches:
        state = "punched_in" if punches[-1] == "in" else "completed"

    summary_out = None
    if summary:
        s = dict(summary)
        if s.get("first_punch_in"):
            s["first_punch_in"] = to_local(s["first_punch_in"]).strftime("%I:%M %p")
        if s.get("last_punch_out"):
            s["last_punch_out"] = to_local(s["last_punch_out"]).strftime("%I:%M %p")
        summary_out = s

    first_punch_in_iso = None
    if "in" in punches and logs:
        for log_entry in logs:
            if log_entry["punch_type"] == "in":
                first_punch_in_iso = log_entry["punched_at"].isoformat()
                break

    return {
        "is_punched_in": state == "punched_in",
        "state": state,
        "last_punch": {"punch_type": last["punch_type"], "punched_at": str(last["punched_at"])} if last else None,
        "summary": summary_out,
        "last_punch_in": first_punch_in_iso,
    }


@app.get("/api/attendance/today")
async def today_logs(
    user: dict = Depends(get_current_user),
    db: asyncpg.Connection = Depends(get_db),
):
    logs = await db.fetch(
        """SELECT punch_type, punched_at, distance_meters FROM attendance_logs
           WHERE user_id=$1
             AND (punched_at AT TIME ZONE $2)::date=(NOW() AT TIME ZONE $2)::date
           ORDER BY punched_at""",
        user["id"], settings.office_timezone,
    )
    return [
        {**dict(r), "punched_at_local": to_local(r["punched_at"]).strftime("%I:%M %p")}
        for r in logs
    ]


# ══════════════════════════════════════════════════════════════
# HR — BRANCHES
# ══════════════════════════════════════════════════════════════

@app.get("/api/hr/branches")
async def get_branches(
    _: dict = Depends(get_current_user),
    db: asyncpg.Connection = Depends(get_db),
):
    rows = await db.fetch(
        "SELECT id, name, city, address, latitude, longitude, radius_meters FROM branches WHERE is_active=TRUE ORDER BY city, name"
    )
    return [{**dict(r), "latitude": float(r["latitude"]), "longitude": float(r["longitude"])} for r in rows]


# ══════════════════════════════════════════════════════════════
# HR — ATTENDANCE REPORTS
# ══════════════════════════════════════════════════════════════

def _ser(e: dict) -> dict:
    """Serialize a report row's time fields."""
    for f in ("first_punch_in", "last_punch_out"):
        if e.get(f):
            e[f] = to_local(e[f]).isoformat()
    for f in ("shift_start", "shift_end"):
        if e.get(f):
            e[f] = e[f].strftime("%H:%M")
    return e

async def regularization_audit(
    employee_id : Optional[int]  = Query(None),
    from_date   : Optional[str]  = Query(None, description="YYYY-MM-DD"),
    to_date     : Optional[str]  = Query(None, description="YYYY-MM-DD"),
    final_status: Optional[str]  = Query(None, description="approved|rejected|pending"),
    page        : int            = Query(1, ge=1),
    page_size   : int            = Query(10, ge=1, le=200),
    _hr         : dict           = Depends(require_hr),
    db          : asyncpg.Connection = Depends(get_db),
):
    """
    HR audit trail for all regularization requests.

    Returns each request with:
      - Full approval chain (L1 + L2 actions with timestamps)
      - Audit log entries (who did what, when, with before/after snapshot)
      - Final daily_summary state

    Performance:
      - Single JOIN query — no N+1
      - Paginated — max 200 rows per call
      - Filtered by indexed columns only (employee_id, work_date, final_status)
    """

    # ── Build filter conditions ───────────────────────────────
    conditions = ["1=1"]
    params: list = []

    def add(val):
        params.append(val)
        return f"${len(params)}"

    if employee_id:
        conditions.append(f"r.employee_id = {add(employee_id)}")
    if from_date:
        try:
            from datetime import datetime as dt
            conditions.append(f"r.work_date >= {add(dt.strptime(from_date, '%Y-%m-%d').date())}")
        except ValueError:
            raise HTTPException(400, "from_date must be YYYY-MM-DD")
    if to_date:
        try:
            from datetime import datetime as dt
            conditions.append(f"r.work_date <= {add(dt.strptime(to_date, '%Y-%m-%d').date())}")
        except ValueError:
            raise HTTPException(400, "to_date must be YYYY-MM-DD")
    if final_status:
        if final_status not in ("approved", "rejected", "pending"):
            raise HTTPException(400, "final_status must be approved|rejected|pending")
        conditions.append(f"r.final_status = {add(final_status)}")

    offset = (page - 1) * page_size
    where  = " AND ".join(conditions)

    # ── Single query — requests + managers + daily_summary ───
    # Audit log entries fetched separately per request (grouped)
    rows = await db.fetch(
        f"""
        SELECT
            r.id                    AS request_id,
            r.work_date,
            r.actual_worked_minutes,
            r.requested_minutes,
            r.reason,
            r.submitted_at,
            r.final_status,

            -- Employee
            emp_u.full_name         AS employee_name,
            emp_e.emp_id,
            emp_e.department,

            -- L1 manager
            r.l1_status,
            r.l1_comment,
            r.l1_approved_at,
            l1u.full_name           AS l1_manager_name,

            -- L2 manager
            r.l2_status,
            r.l2_comment,
            r.l2_approved_at,
            l2u.full_name           AS l2_manager_name,

            -- Current daily_summary state (what employee actually sees)
            COALESCE(ds.total_minutes,   0) AS current_total_minutes,
            COALESCE(ds.payroll_status, 'absent') AS current_payroll_status,
            ds.is_regularized,
            ds.payroll_notes

        FROM regularization_requests r
        JOIN employees emp_e ON emp_e.id = r.employee_id
        JOIN users     emp_u ON emp_u.id = emp_e.user_id
        LEFT JOIN employees l1e ON l1e.id = r.l1_manager_id
        LEFT JOIN users     l1u ON l1u.id = l1e.user_id
        LEFT JOIN employees l2e ON l2e.id = r.l2_manager_id
        LEFT JOIN users     l2u ON l2u.id = l2e.user_id
        LEFT JOIN daily_summary ds
               ON ds.user_id = emp_e.user_id AND ds.work_date = r.work_date
        WHERE {where}
        ORDER BY r.work_date DESC, r.submitted_at DESC
        LIMIT {add(page_size)} OFFSET {add(offset)}
        """,
        *params,
    )

    if not rows:
        return {"total": 0, "stats": {"approved": 0, "rejected": 0, "pending": 0}, "page": page, "page_size": page_size, "requests": []}

    # ── Fetch audit log entries for all returned requests ─────
    # One query for all request_ids — no N+1
    request_ids = [r["request_id"] for r in rows]
    audit_rows  = await db.fetch(
        """
        SELECT
            al.request_id,
            al.action_role,
            al.action_type,
            al.note,
            al.minutes_before,
            al.payroll_status_before,
            al.minutes_after,
            al.payroll_status_after,
            al.created_at,
            u.full_name AS actioned_by
        FROM regularization_audit_logs al
        LEFT JOIN users u ON u.id = al.action_by_user_id
        WHERE al.request_id = ANY($1::int[])
        ORDER BY al.created_at ASC
        """,
        request_ids,
    )

    # Group audit entries by request_id
    audit_map: dict[int, list] = {}
    for a in audit_rows:
        rid = a["request_id"]
        if rid not in audit_map:
            audit_map[rid] = []
        audit_map[rid].append({
            "action_role"          : a["action_role"],
            "action_type"          : a["action_type"],
            "actioned_by"          : a["actioned_by"],
            "note"                 : a["note"],
            "minutes_before"       : a["minutes_before"],
            "payroll_status_before": a["payroll_status_before"],
            "minutes_after"        : a["minutes_after"],
            "payroll_status_after" : a["payroll_status_after"],
            "created_at"           : a["created_at"].isoformat(),
        })

    # ── Count total + status breakdown for pagination KPIs ─────
    # params[:-2] strips page_size and offset added for LIMIT/OFFSET
    count_params = params[:-2]
    total_row = await db.fetchrow(
        f"SELECT COUNT(*) AS total FROM regularization_requests r WHERE {where}",
        *count_params,
    )
    # One extra query for KPI counts — reuses same WHERE, no N+1
    stats_rows = await db.fetch(
        f"SELECT final_status, COUNT(*) AS cnt FROM regularization_requests r WHERE {where} GROUP BY final_status",
        *count_params,
    )
    stats = {r["final_status"]: r["cnt"] for r in stats_rows}

    # ── Build response ────────────────────────────────────────
    def fmt_min(m: int) -> str:
        if not m:
            return "0m"
        h, mn = divmod(m, 60)
        return f"{h}h {mn}m" if h else f"{mn}m"

    result = []
    for r in rows:
        result.append({
            "request_id"            : r["request_id"],
            "work_date"             : r["work_date"].isoformat(),
            "employee_name"         : r["employee_name"],
            "emp_id"                : r["emp_id"],
            "department"            : r["department"],

            "actual_worked"         : fmt_min(r["actual_worked_minutes"]),
            "requested"             : fmt_min(r["requested_minutes"]),
            "reason"                : r["reason"],
            "submitted_at"          : r["submitted_at"].isoformat(),
            "final_status"          : r["final_status"],

            "l1_manager"            : r["l1_manager_name"],
            "l1_status"             : r["l1_status"],
            "l1_comment"            : r["l1_comment"],
            "l1_approved_at"        : r["l1_approved_at"].isoformat() if r["l1_approved_at"] else None,

            "l2_manager"            : r["l2_manager_name"],
            "l2_status"             : r["l2_status"],
            "l2_comment"            : r["l2_comment"],
            "l2_approved_at"        : r["l2_approved_at"].isoformat() if r["l2_approved_at"] else None,

            # What the employee's record currently shows
            "current_total_minutes" : r["current_total_minutes"],
            "current_total_display" : fmt_min(r["current_total_minutes"]),
            "current_payroll_status": r["current_payroll_status"],
            "is_regularized"        : r["is_regularized"],
            "payroll_notes"         : r["payroll_notes"],

            # Full audit trail for this request
            "audit_trail"           : audit_map.get(r["request_id"], []),
        })

    return {
        "total"    : total_row["total"],
        "stats"    : {
            "approved": stats.get("approved", 0),
            "rejected": stats.get("rejected", 0),
            "pending" : stats.get("pending",  0),
        },
        "page"     : page,
        "page_size": page_size,
        "requests" : result,
    }


@app.get("/api/hr/regularization-audit")
async def hr_regularization_audit(
    employee_id : Optional[int]  = Query(None),
    from_date   : Optional[str]  = Query(None),
    to_date     : Optional[str]  = Query(None),
    final_status: Optional[str]  = Query(None),
    page        : int            = Query(1, ge=1),
    page_size   : int            = Query(50, ge=1, le=200),
    hr          : dict           = Depends(require_hr),
    db          : asyncpg.Connection = Depends(get_db),
):
    return await regularization_audit(
        employee_id=employee_id, from_date=from_date, to_date=to_date,
        final_status=final_status, page=page, page_size=page_size,
        _hr=hr, db=db,
    )

@app.get("/api/hr/daily-report")
async def daily_report(
    date_str:  Annotated[Optional[str], Query()] = None,
    branch_id: Annotated[Optional[int], Query()] = None,
    _hr: dict = Depends(require_hr),
    db: asyncpg.Connection = Depends(get_db),
):
    target = parse_date_param(date_str)
    q = """
        SELECT u.id, u.email, u.full_name,
            e.shift_start, e.shift_end,
            b.name AS branch_name, b.city,
            s.first_punch_in, s.last_punch_out, s.total_minutes,
            s.is_late, s.late_by_minutes,
            COALESCE(s.status,         'absent') AS status,
            COALESCE(s.payroll_status, 'absent') AS payroll_status
        FROM users u
        JOIN employees e ON e.user_id = u.id
        LEFT JOIN branches b ON b.id = e.branch_id
        LEFT JOIN daily_summary s ON s.user_id=u.id AND s.work_date=$1
        WHERE u.is_active=TRUE AND e.is_active=TRUE AND u.role='employee'
    """
    params: list = [target]
    if branch_id:
        q += " AND e.branch_id=$2"
        params.append(branch_id)
    q += " ORDER BY b.name NULLS LAST, u.full_name"

    rows = await db.fetch(q, *params)
    employees = [_ser(dict(r)) for r in rows]
    total = len(employees)
    return {
        "date": target.isoformat(),
        "stats": {
            "total":   total,
            "present": sum(1 for e in employees if e["payroll_status"] == "present"),
            "absent":  sum(1 for e in employees if e["payroll_status"] == "absent"),
            "on_leave": sum(1 for e in employees if e["status"] == "leave"),
            "late":    sum(1 for e in employees if e.get("is_late")),
        },
        "employees": employees,
    }


@app.get("/api/hr/export")
async def export_excel(
    date_str:  Annotated[str, Query()],
    branch_id: Annotated[Optional[int], Query()] = None,
    _hr: dict = Depends(require_hr),
    db: asyncpg.Connection = Depends(get_db),
):
    target = parse_date_param(date_str)
    q = """
        SELECT u.email, u.full_name, b.name AS branch_name, b.city,
            s.first_punch_in, s.last_punch_out, s.total_minutes,
            s.is_late, s.late_by_minutes,
            COALESCE(s.status,         'absent') AS attendance_status,
            COALESCE(s.payroll_status, 'absent') AS payroll_status
        FROM users u
        JOIN employees e ON e.user_id = u.id
        LEFT JOIN branches b ON b.id = e.branch_id
        LEFT JOIN daily_summary s ON s.user_id=u.id AND s.work_date=$1
        WHERE u.is_active=TRUE AND e.is_active=TRUE AND u.role='employee'
    """
    params: list = [target]
    if branch_id:
        q += " AND e.branch_id=$2"
        params.append(branch_id)
    q += " ORDER BY b.name NULLS LAST, u.full_name"
    rows = await db.fetch(q, *params)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    headers = ["Email", "Name", "Branch", "City", "Punch In", "Punch Out", "Hours", "Attendance", "Payroll", "Late By"]
    hfill = PatternFill(start_color="3B63F6", end_color="3B63F6", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h); c.font = hfont; c.fill = hfill; c.alignment = Alignment(horizontal="center")

    FILL_GREEN  = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
    FILL_BLUE   = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")  # leave
    FILL_RED    = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")

    for ri, row in enumerate(rows, 2):
        att    = row["attendance_status"]
        pay    = row["payroll_status"]
        # Row colour driven by payroll: green=paid, blue=leave(paid), red=absent
        if att == "leave" and pay == "present":
            fill = FILL_BLUE
        elif pay == "present":
            fill = FILL_GREEN
        else:
            fill = FILL_RED
        mins = row["total_minutes"] or 0
        late_prefix = "LATE — " if row["is_late"] else ""
        vals = [
            row["email"], row["full_name"],
            row["branch_name"] or "—", row["city"] or "—",
            to_local(row["first_punch_in"]).strftime("%I:%M %p") if row["first_punch_in"] else "—",
            to_local(row["last_punch_out"]).strftime("%I:%M %p") if row["last_punch_out"] else "—",
            f"{mins//60}h {mins%60}m" if mins else "—",
            late_prefix + att.upper(),    # attendance: PRESENT / LEAVE / ABSENT
            pay.upper(),                  # payroll: PRESENT / ABSENT
            f"+{row['late_by_minutes']}m" if row["is_late"] else "On Time",
        ]
        for col, val in enumerate(vals, 1):
            ws.cell(ri, col, val).fill = fill

    for col, w in zip("ABCDEFGHIJ", [26, 22, 26, 16, 12, 12, 10, 14, 10, 10]):
        ws.column_dimensions[col].width = w

    out = BytesIO(); wb.save(out); out.seek(0)
    return StreamingResponse(out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="attendance_{target}.xlsx"'})

# ══════════════════════════════════════════════════════════════
# HR — ONBOARDING / EMPLOYEES
# ══════════════════════════════════════════════════════════════

@app.post("/api/hr/employees")
async def onboard_employee(
    payload: OnboardRequest = Body(...),
    hr: dict = Depends(require_hr),
    db: asyncpg.Connection = Depends(get_db),
):
    """
    HR onboards a new employee (atomic operation):
    - Creates user (auth identity)
    - Creates employee profile (1:1)
    - Generates final password (usable directly)
    - Logs credential generation audit
    
    Requires: HR or Admin authentication
    """

    validate_onboard_payload(payload)

    # Guard: L1 and L2 cannot be the same person
    if (payload.l1_manager_id and payload.l2_manager_id
            and payload.l1_manager_id == payload.l2_manager_id):
        raise HTTPException(400, "L1 and L2 manager cannot be the same person")

    # 1. Validate role (only employee role allowed here)
    if payload.role not in ("employee", "hr"):
        raise HTTPException(400, "Only 'employee' or 'hr' roles can be onboarded here")

    # 2. Require personal_email — it is the login identifier
    if not payload.personal_email:
        raise HTTPException(400, "Personal email is required (used for login and credential delivery)")

    # Check duplicate work email in users
    existing = await db.fetchrow(
        "SELECT id FROM users WHERE LOWER(email) = LOWER($1)",
        payload.work_email,
    )
    if existing:
        raise HTTPException(409, "User with this work email already exists")

    # Check duplicate personal_email in employees — must be unique for login
    dup_personal = await db.fetchrow(
        "SELECT id FROM employees WHERE LOWER(personal_email) = LOWER($1)",
        payload.personal_email,
    )
    if dup_personal:
        raise HTTPException(409, "An employee with this personal email already exists")

    # 3. Generate SDPL emp_id — find highest existing number and increment
    last = await db.fetchrow(
        """SELECT emp_id FROM employees
           WHERE emp_id ~ '^SDPL[0-9]+$'
           ORDER BY CAST(SUBSTRING(emp_id FROM 5) AS INTEGER) DESC
           LIMIT 1"""
    )
    if last and last["emp_id"]:
        next_num = int(last["emp_id"][4:]) + 1
    else:
        next_num = 1
    new_emp_id = f"SDPL{next_num:03d}"

    # 4. Generate final password
    raw_password = payload.password or generate_temp_password()
    hashed_password = pwd_context.hash(raw_password)

    async with db.transaction():
        # 4. Create user
        user = await db.fetchrow(
            """
            INSERT INTO users (email, password_hash, full_name, role, is_active, must_reset_password)
            VALUES ($1, $2, $3, $4, TRUE, TRUE)
            RETURNING id, email, full_name, role
            """,
            payload.work_email,
            hashed_password,
            payload.full_name,
            payload.role,
        )
        user_id = user["id"]

        # 5. Create employee profile (mandatory for employee/hr roles)
        emp = await db.fetchrow(
            """
            INSERT INTO employees (
                user_id, emp_id,
                phone, personal_email, dob, gender, blood_group, nationality,
                home_address,
                branch_id, job_title, designation, department, sub_department,
                grade, date_of_joining, cost_centre,
                l1_manager_id, l2_manager_id,
                employment_type, contract_end, probation_end, notice_period,
                shift_start, shift_end, work_mode, weekly_off, work_location, asset_id,
                annual_ctc, pay_frequency, pf_enrolled, esic_applicable,
                bank_name, bank_account, bank_ifsc, pan_number,
                emg_name, emg_phone, emg_rel,
                onboarding_status
            )
            VALUES (
                $1,$2,
                $3,$4,$5,$6,$7,$8,
                $9,
                $10,$11,$12,$13,$14,
                $15,$16,$17,
                $18,$19,
                $20,$21,$22,$23,
                $24,$25,$26,$27,$28,$29,
                $30,$31,$32,$33,
                $34,$35,$36,$37,
                $38,$39,$40,
                'completed'
            )
            RETURNING id
            """,
            user_id, new_emp_id,
            payload.phone, payload.personal_email, parse_date(payload.dob), payload.gender,
            payload.blood_group, payload.nationality,
            payload.home_address,
            payload.branch_id, payload.job_title, payload.designation,
            payload.department, payload.sub_department,
            payload.grade, parse_date(payload.date_of_joining), payload.cost_centre,
            payload.l1_manager_id, payload.l2_manager_id,
            payload.employment_type, parse_date(payload.contract_end), parse_date(payload.probation_end), payload.notice_period,
            payload.shift_start, payload.shift_end, payload.work_mode,
            payload.weekly_off, payload.work_location, payload.asset_id,
            payload.annual_ctc, payload.pay_frequency, payload.pf_enrolled,
            payload.esic_applicable,
            payload.bank_name, payload.bank_account, payload.bank_ifsc, payload.pan_number,
            payload.emg_name, payload.emg_phone, payload.emg_rel,
        )

        employee_id = emp["id"]

        # 6. Credential audit entry
        await db.execute(
            """
            INSERT INTO credential_audits
            (user_id, action, performed_by, is_temporary, notes, created_at)
            VALUES ($1, 'generated', $2, FALSE, $3, NOW())
            """,
            user_id,
            hr["id"],
            f"Employee onboarded by HR ({hr['email']})",
        )

    logger.info(
        "Employee onboarded: emp_id=%s user_id=%s by hr_id=%s",
        employee_id, user_id, hr["id"]
    )

    # Send credentials to personal email (login identifier — not work email)
    email_sent = send_welcome_credentials(payload.personal_email, user["full_name"], raw_password)

    return {
        "employee_id": employee_id,
        "emp_id": new_emp_id,
        "email": user["email"],               # work email
        "personal_email": payload.personal_email,  # login email — shown in credentials modal
        "full_name": user["full_name"],
        "role": user["role"],
        "temporary_password": raw_password,
        "email_sent": email_sent,
        "message": "Employee onboarded successfully with credentials.",
        "created_at": datetime.now().isoformat(),
    }

@app.get("/api/hr/employees")
async def list_employees(
    search:            Annotated[Optional[str], Query()] = None,
    department:        Annotated[Optional[str], Query()] = None,
    branch_id:         Annotated[Optional[int], Query()] = None,
    onboarding_status: Annotated[Optional[str], Query()] = None,
    is_active: Annotated[Optional[bool], Query()] = None,
    _hr: dict = Depends(require_hr),
    db: asyncpg.Connection = Depends(get_db),
):
    """List employees with manager names. Filterable."""
    # conditions = ["u.is_active = TRUE", "e.is_active = TRUE"]
    conditions = ["1=1"]  # for easier appending of AND clauses
    params: list = []

    def add(val):
        params.append(val); return f"${len(params)}"

    if search:
        ph = add(f"%{search}%")
        conditions.append(f"(u.full_name ILIKE {ph} OR u.email ILIKE {ph} OR e.emp_id ILIKE {ph})")
    if department:
        conditions.append(f"e.department = {add(department)}")
    if branch_id:
        conditions.append(f"e.branch_id = {add(branch_id)}")
    if onboarding_status:
        conditions.append(f"e.onboarding_status = {add(onboarding_status)}")
    if is_active is not None:
        conditions.append(f"e.is_active = {add(is_active)}")

    rows = await db.fetch(
        f"""SELECT
              u.id AS user_id, e.id, e.emp_id,
              u.email, u.full_name, u.role,
              e.is_active AS is_active,
              e.phone, e.job_title, e.designation, e.department,
              e.grade, e.date_of_joining, e.onboarding_status,
              e.work_mode, e.shift_start, e.shift_end,
              e.l1_manager_id, e.l2_manager_id,
              l1u.full_name AS l1_name, l1e.job_title AS l1_title,
              l2u.full_name AS l2_name, l2e.job_title AS l2_title,
              b.name AS branch_name, b.city AS branch_city,
              e.created_at
            FROM users u
            JOIN employees e   ON e.user_id = u.id
            LEFT JOIN branches b  ON b.id = e.branch_id
            LEFT JOIN employees l1e ON l1e.id = e.l1_manager_id
            LEFT JOIN users     l1u ON l1u.id = l1e.user_id
            LEFT JOIN employees l2e ON l2e.id = e.l2_manager_id
            LEFT JOIN users     l2u ON l2u.id = l2e.user_id
            WHERE {' AND '.join(conditions)}
            ORDER BY e.created_at DESC""",
        *params,
    )

    result = []
    for r in rows:
        d = dict(r)
        for f in ("shift_start", "shift_end"):
            if d.get(f): d[f] = d[f].strftime("%H:%M")
        if d.get("date_of_joining"): d["date_of_joining"] = d["date_of_joining"].isoformat()
        if d.get("created_at"):      d["created_at"] = d["created_at"].isoformat()
        result.append(d)

    return {"total": len(result), "employees": result}


@app.get("/api/hr/employees/{emp_id}")
async def get_employee(
    emp_id: int,
    _hr: dict = Depends(require_hr),
    db: asyncpg.Connection = Depends(get_db),
):
    """Full detail for one employee."""
    row = await db.fetchrow(
        """SELECT
              u.id AS user_id, e.id, e.emp_id,
              u.email, u.full_name, u.role, u.last_login,
              e.phone, e.personal_email, e.dob, e.gender, e.blood_group,
              e.nationality, e.home_address,
              e.emg_name, e.emg_phone, e.emg_rel,
              e.branch_id, e.job_title, e.designation, e.department,
              e.sub_department, e.grade, e.date_of_joining, e.cost_centre,
              e.l1_manager_id, e.l2_manager_id,
              l1u.full_name AS l1_name, l1e.job_title AS l1_title, l1u.role AS l1_role,
              l2u.full_name AS l2_name, l2e.job_title AS l2_title, l2u.role AS l2_role,
              e.employment_type, e.contract_end, e.probation_end, e.notice_period,
              e.shift_start, e.shift_end, e.work_mode, e.weekly_off,
              e.work_location, e.asset_id,
              e.annual_ctc, e.pay_frequency, e.pf_enrolled, e.esic_applicable,
              e.bank_name, e.bank_account, e.bank_ifsc, e.pan_number,
              e.onboarding_status, e.created_at, e.updated_at,
              b.name AS branch_name, b.city AS branch_city
           FROM employees e
           JOIN users u ON u.id = e.user_id
           LEFT JOIN branches b    ON b.id = e.branch_id
           LEFT JOIN employees l1e ON l1e.id = e.l1_manager_id
           LEFT JOIN users     l1u ON l1u.id = l1e.user_id
           LEFT JOIN employees l2e ON l2e.id = e.l2_manager_id
           LEFT JOIN users     l2u ON l2u.id = l2e.user_id
           WHERE e.id = $1""",
        emp_id,
    )
    if not row:
        raise HTTPException(404, "Employee not found")

    d = dict(row)
    for f in ("shift_start", "shift_end"):
        if d.get(f): d[f] = d[f].strftime("%H:%M")
    for f in ("dob", "date_of_joining", "contract_end", "probation_end"):
        if d.get(f): d[f] = d[f].isoformat()
    for f in ("created_at", "updated_at", "last_login"):
        if d.get(f): d[f] = d[f].isoformat()
    if d.get("annual_ctc"): d["annual_ctc"] = float(d["annual_ctc"])
    return d

# from fastapi import Path as PathParam

@app.post("/api/hr/employees/{emp_id}/generate-credentials")
async def regenerate_employee_credentials(
    emp_id: int = PathParam(..., gt=0),
    hr: dict = Depends(require_hr),
    db: asyncpg.Connection = Depends(get_db),
):
    """
    HR regenerates final credentials for an employee.
    - Overwrites password_hash
    - Does NOT force password reset (simplified lifecycle)
    - Returns new password one-time to HR
    """

    # 1. Fetch employee + user info
    emp = await db.fetchrow(
        """
        SELECT e.id, e.user_id, u.email, u.full_name, u.is_active,
               e.personal_email
        FROM employees e
        JOIN users u ON u.id = e.user_id
        WHERE e.id = $1
        """,
        emp_id,
    )

    if not emp:
        raise HTTPException(404, "Employee not found")

    if not emp["is_active"]:
        raise HTTPException(400, "Cannot generate credentials for inactive employee")

    user_id = emp["user_id"]

    # 2. Generate final password
    new_password = generate_temp_password()
    hashed = pwd_context.hash(new_password)

    async with db.transaction():
        # 3. Update password (final credential, no reset enforcement)
        await db.execute(
            """
            UPDATE users
            SET password_hash=$1, must_reset_password=TRUE, last_login=NULL
            WHERE id=$2
            """,
            hashed,
            user_id,
        )

        # 4. Insert audit log
        await db.execute(
            """
            INSERT INTO credential_audits
            (user_id, action, performed_by, is_temporary, notes, created_at)
            VALUES ($1, 'generated_final', $2, FALSE, $3, NOW())
            """,
            user_id,
            hr["id"],
            f"Final credentials regenerated by HR ({hr['email']})",
        )

    # Send new credentials to personal email (login identifier)
    send_to = emp["personal_email"] or emp["email"]  # fallback to work email if personal not set
    email_sent = send_welcome_credentials(send_to, emp["full_name"], new_password)

    return {
        "employee_id": emp_id,
        "email": emp["email"],
        "full_name": emp["full_name"],
        "new_password": new_password,  # return only once
        "email_sent": email_sent,
        "message": "New credentials generated successfully.",
        "generated_at": datetime.now().isoformat(),
    }

@app.patch("/api/hr/employees/{emp_id}/onboarding-status")
async def update_onboarding_status(
    emp_id: int,
    req: UpdateOnboardingStatus,
    hr: dict = Depends(require_hr),
    db: asyncpg.Connection = Depends(get_db),
):
    result = await db.execute(
        "UPDATE employees SET onboarding_status=$1, updated_at=NOW() WHERE id=$2",
        req.status, emp_id,
    )
    if result == "UPDATE 0":
        raise HTTPException(404, "Employee not found")
    logger.info("Onboarding status: emp_id=%s → %s by hr=%s", emp_id, req.status, hr["id"])
    return {"id": emp_id, "onboarding_status": req.status}



@app.put("/api/hr/employees/{emp_id}")
async def update_employee(
    emp_id: int,
    req: UpdateEmployeeRequest,
    hr: dict = Depends(require_hr),
    db: asyncpg.Connection = Depends(get_db),
):
    """Update employee fields (partial). Only non-None fields are written."""
    validate_update_payload(req)

    # Guard: L1 and L2 cannot be the same person
    if (req.l1_manager_id and req.l2_manager_id
            and req.l1_manager_id == req.l2_manager_id):
        raise HTTPException(400, "L1 and L2 manager cannot be the same person")

    # Guard: employee cannot be their own manager
    if req.l1_manager_id and req.l1_manager_id == emp_id:
        raise HTTPException(400, "Employee cannot be their own L1 manager")
    if req.l2_manager_id and req.l2_manager_id == emp_id:
        raise HTTPException(400, "Employee cannot be their own L2 manager")

    row = await db.fetchrow("SELECT e.id, e.user_id FROM employees e WHERE e.id = $1", emp_id)
    if not row:
        raise HTTPException(404, "Employee not found")
    user_id = row["user_id"]

    emp_fields = []
    simple_emp = [
        ("personal_email", req.personal_email), ("phone", req.phone),
        ("gender", req.gender), ("blood_group", req.blood_group),
        ("nationality", req.nationality), ("home_address", req.home_address),
        ("emg_name", req.emg_name), ("emg_phone", req.emg_phone), ("emg_rel", req.emg_rel),
        ("job_title", req.job_title), ("designation", req.designation),
        ("department", req.department), ("sub_department", req.sub_department),
        ("grade", req.grade), ("cost_centre", req.cost_centre),
        ("employment_type", req.employment_type), ("notice_period", req.notice_period),
        ("work_mode", req.work_mode), ("weekly_off", req.weekly_off),
        ("work_location", req.work_location), ("asset_id", req.asset_id),
        ("pay_frequency", req.pay_frequency), ("bank_name", req.bank_name),
        ("bank_account", req.bank_account), ("bank_ifsc", req.bank_ifsc),
        ("pan_number", req.pan_number),
    ]
    for col, val in simple_emp:
        if val is not None:
            emp_fields.append((col, val))

    if req.branch_id is not None:       emp_fields.append(("branch_id",       req.branch_id or None))
    if req.l1_manager_id is not None:   emp_fields.append(("l1_manager_id",   req.l1_manager_id or None))
    if req.l2_manager_id is not None:   emp_fields.append(("l2_manager_id",   req.l2_manager_id or None))
    if req.annual_ctc is not None:      emp_fields.append(("annual_ctc",      req.annual_ctc))
    if req.pf_enrolled is not None:     emp_fields.append(("pf_enrolled",     req.pf_enrolled))
    if req.esic_applicable is not None: emp_fields.append(("esic_applicable", req.esic_applicable))
    if req.dob is not None:             emp_fields.append(("dob",             parse_date(req.dob)))
    if req.date_of_joining is not None: emp_fields.append(("date_of_joining", parse_date(req.date_of_joining)))
    if req.contract_end is not None:    emp_fields.append(("contract_end",    parse_date(req.contract_end)))
    if req.probation_end is not None:   emp_fields.append(("probation_end",   parse_date(req.probation_end)))
    if req.shift_start is not None:
        try:    emp_fields.append(("shift_start", time.fromisoformat(req.shift_start)))
        except ValueError: raise HTTPException(400, "Invalid shift_start, use HH:MM")
    if req.shift_end is not None:
        try:    emp_fields.append(("shift_end", time.fromisoformat(req.shift_end)))
        except ValueError: raise HTTPException(400, "Invalid shift_end, use HH:MM")

    async with db.transaction():
        if emp_fields:
            params = [v for _, v in emp_fields]
            sets = ", ".join(f"{col}=${i+1}" for i, (col, _) in enumerate(emp_fields))
            params.append(emp_id)
            await db.execute(
                f"UPDATE employees SET {sets}, updated_at=NOW() WHERE id=${len(params)}", *params)
        if req.full_name is not None:
            await db.execute("UPDATE users SET full_name=$1 WHERE id=$2", req.full_name, user_id)
        if req.role is not None:
            await db.execute("UPDATE users SET role=$1 WHERE id=$2", req.role, user_id)

    logger.info("Employee updated: emp_id=%s by hr=%s", emp_id, hr["id"])
    return {"id": emp_id, "message": "Employee updated successfully"}


@app.patch("/api/hr/employees/{emp_id}/deactivate")
async def deactivate_employee(
    emp_id: int = PathParam(..., gt=0),
    hr: dict = Depends(require_hr),
    db: asyncpg.Connection = Depends(get_db),
):
    """
    Soft delete an employee:
    - Marks employee inactive
    - Disables login by setting users.is_active = FALSE
    - Keeps all historical data (attendance, audits)
    """

    # 1. Fetch employee + user status
    emp = await db.fetchrow(
        """
        SELECT e.id, e.user_id, e.is_active AS emp_active,
               u.email, u.full_name, u.is_active AS user_active
        FROM employees e
        JOIN users u ON u.id = e.user_id
        WHERE e.id = $1
        """,
        emp_id,
    )

    if not emp:
        raise HTTPException(404, "Employee not found")

    if not emp["emp_active"]:
        raise HTTPException(400, "Employee already inactive")

    async with db.transaction():
        # 2. Deactivate employee profile
        await db.execute(
            "UPDATE employees SET is_active=FALSE WHERE id=$1",
            emp_id,
        )

        # 3. Disable login
        await db.execute(
            "UPDATE users SET is_active=FALSE WHERE id=$1",
            emp["user_id"],
        )

        # # 4. Audit log
        # await db.execute(
        #     """
        #     INSERT INTO credential_audits
        #     (user_id, action, performed_by, is_temporary, notes, created_at)
        #     VALUES ($1, 'deactivated', $2, FALSE, $3, NOW())
        #     """,
        #     emp["user_id"],
        #     hr["id"],
        #     f"Employee deactivated by HR ({hr['email']})",
        # )

    return {
        "employee_id": emp_id,
        "email": emp["email"],
        "full_name": emp["full_name"],
        "status": "inactive",
        "message": "Employee soft-deactivated successfully. Login disabled.",
    }

@app.patch("/api/hr/employees/{emp_id}/reactivate")
async def reactivate_employee(
    emp_id: int,
    hr: dict = Depends(require_hr),
    db: asyncpg.Connection = Depends(get_db),
):
    """Reactivate a deactivated employee."""
    
    # 1. Fetch employee + user
    row = await db.fetchrow(
        "SELECT id, user_id, is_active FROM employees WHERE id = $1",
        emp_id,
    )
    if not row:
        raise HTTPException(404, "Employee not found")

    if row["is_active"]:
        raise HTTPException(400, "Employee is already active")

    # 2. Reactivate both employee profile and user login
    async with db.transaction():
        await db.execute(
            "UPDATE employees SET is_active=TRUE WHERE id=$1",
            emp_id,
        )
        await db.execute(
            "UPDATE users SET is_active=TRUE WHERE id=$1",
            row["user_id"],
        )

        # # Optional audit log
        # await db.execute(
        #     """
        #     INSERT INTO credential_audits
        #     (user_id, action, performed_by, is_temporary, notes, created_at)
        #     VALUES ($1, 'reactivated', $2, FALSE, $3, NOW())
        #     """,
        #     row["user_id"],
        #     hr["id"],
        #     f"Employee reactivated by HR ({hr['email']})",
        # )

    logger.info("Employee reactivated: emp_id=%s by hr=%s", emp_id, hr["id"])

    return {
        "id": emp_id,
        "message": "Employee reactivated successfully",
        "status": "active",
    }


@app.get("/api/hr/onboarding-stats")
async def onboarding_stats(
    _hr: dict = Depends(require_hr),
    db: asyncpg.Connection = Depends(get_db),
):
    row = await db.fetchrow(
        """SELECT
              COUNT(*)                                                    AS total,
              COUNT(*) FILTER (WHERE onboarding_status='awaiting')       AS awaiting,
              COUNT(*) FILTER (WHERE onboarding_status='in-progress')    AS in_progress,
              COUNT(*) FILTER (WHERE onboarding_status='completed')      AS completed
           FROM employees
           WHERE is_active = TRUE"""
    )
    return dict(row)


@app.get("/api/hr/managers")
async def list_managers(
    _: dict = Depends(get_current_user),
    db: asyncpg.Connection = Depends(get_db),
):
    """
    All active employees eligible as L1 or L2 manager.
    Any role qualifies — employee, hr, or admin.
    """
    rows = await db.fetch(
        """SELECT e.id, e.emp_id, u.full_name, u.role,
                  e.job_title, e.designation, e.department
           FROM employees e
           JOIN users u ON u.id = e.user_id
           WHERE u.is_active = TRUE AND e.is_active = TRUE
           ORDER BY u.full_name""",
    )
    return [dict(r) for r in rows]


# ══════════════════════════════════════════════════════════════
# HEALTH + STATIC
# ══════════════════════════════════════════════════════════════

@app.get("/health")
async def health():
    tz = pytz.timezone(settings.office_timezone)
    return {
        "status": "healthy" if db_pool else "degraded",
        "utc":    datetime.now(pytz.utc).isoformat(),
        "local":  datetime.now(tz).isoformat(),
    }


BASE_DIR = Path(__file__).resolve().parent
FRONTEND_DIR = BASE_DIR.parent / "frontend"


@app.get("/login")
async def login_page():
    return FileResponse(FRONTEND_DIR / "index.html")


@app.get("/employee")
async def employee_portal():
    return FileResponse(FRONTEND_DIR / "employee.html")


@app.get("/hr-manager")
async def hr_manager_portal():
    return FileResponse(FRONTEND_DIR / "hr.html")

@app.get("/regularization")
async def regularization_page():
    return FileResponse(FRONTEND_DIR / "regularization.html")

@app.get("/leave")
async def leave_page():
    return FileResponse(FRONTEND_DIR / "leave.html")

@app.get("/holidays")
async def holidays_page():
    return FileResponse(FRONTEND_DIR / "holidays.html")

@app.get("/payroll")
async def payroll_page():
    return FileResponse(FRONTEND_DIR / "payroll.html")


app.mount("/", StaticFiles(directory=str(FRONTEND_DIR), html=True), name="frontend")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=False)