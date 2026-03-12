"""
payroll.py — Payroll Engine

Flow:
  1. HR picks month + branch/employees
  2. GET /api/payroll/export  → pulls DB data, returns pre-filled Excel
  3. HR reviews, corrects if needed
  4. POST /api/payroll/process → accepts corrected Excel, calculates, returns:
       - Excel summary (all employees)
       - ZIP of individual PDF payslips

Salary formula (all derived from annual_ctc stored in employees table):
  fixed_monthly = annual_ctc / 12
  basic         = fixed_monthly * 0.40
  hra           = basic * 0.50
  ca            = basic * 0.15
  sa            = fixed_monthly - (basic + hra + ca)
  gross         = fixed_monthly  (always)

Deductions:
  lop     = gross / calendar_days * absent_days
  pf      = min(basic * 0.12, 1800)  [if pf_enrolled]
  esi     = TBD                       [if esic_applicable AND gross <= 15000]
  pt      = TBD
  net_pay = gross - lop - pf - esi - pt

Attendance source: Excel wins over DB on conflict (HR's correction is final).

Locking: Once payroll is processed and payslips generated, it is final.
         No re-run endpoint — HR must export again for a new month.
"""

import io
import logging
import calendar
import zipfile
from datetime import date, timedelta
from decimal import Decimal, ROUND_HALF_UP
from typing import Optional

import asyncpg
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from auth import require_hr
from db import get_db

logger = logging.getLogger("payroll")
router = APIRouter(prefix="/api/payroll", tags=["payroll"])


# ══════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════

PF_RATE        = Decimal("0.12")
PF_CAP         = Decimal("1800")    # max PF deduction per month
ESI_THRESHOLD  = Decimal("21000")   # gross <= this → ESI applicable (updated 2017 threshold)
ESI_EMP_RATE   = Decimal("0.0075")  # employee contribution: 0.75%

# Professional Tax slabs (Andhra Pradesh / Telangana)
# Applied on gross salary per month. Slab: (up_to, pt_amount)
# gross > 20000 → ₹200, 15001–20000 → ₹150, 10001–15000 → ₹110, ≤10000 → ₹0
PT_SLABS = [
    (Decimal("10000"),  Decimal("0")),
    (Decimal("15000"),  Decimal("110")),
    (Decimal("20000"),  Decimal("150")),
    (None,              Decimal("200")),   # None = no upper bound
]

# Excel column order — MUST stay in sync between export and process
COLUMNS = [
    "emp_id",
    "employee_name",
    "designation",
    "department",
    "branch",
    "date_of_joining",
    "bank_name",
    "bank_account",
    "bank_ifsc",
    "pan_number",
    "uan_number",
    "pf_enrolled",
    "esic_applicable",
    "annual_ctc",
    "fixed_monthly",
    "basic",
    "hra",
    "ca",
    "sa",
    "gross",
    "per_day_salary",
    "calendar_days",
    "present_days",
    "leaves_taken",
    "lop_days",
    "available_leaves",
    "carry_forward_leaves",
    "lop",
    "pf_deduction",
    "pt_deduction",
    "esi_deduction",
    "net_pay",
]

# Human-readable headers for the Excel sheet
HEADERS = [
    "Emp ID",
    "Employee Name",
    "Designation",
    "Department",
    "Branch",
    "Date of Joining",
    "Bank Name",
    "Bank Account",
    "Bank IFSC",
    "PAN",
    "UAN",
    "PF Enrolled",
    "ESI Applicable",
    "Annual CTC",
    "Fixed Monthly",
    "Basic",
    "HRA",
    "Conveyance",
    "Special Allowance",
    "Gross",
    "Per Day Salary",
    "Calendar Days",
    "Present Days",
    "Leaves Taken",
    "LOP Days",
    "Available Leaves",
    "Carry Forward Leaves",
    "LOP Deduction",
    "PF Deduction",
    "PT Deduction",
    "ESI Deduction",
    "Net Pay",
]

# Columns HR is allowed to edit (zero-indexed positions in COLUMNS list)
# Everything else is protected/informational
EDITABLE_COLS = {
    COLUMNS.index("present_days"),
    COLUMNS.index("lop_days"),
    COLUMNS.index("lop"),
    COLUMNS.index("pf_deduction"),
    COLUMNS.index("pt_deduction"),
    COLUMNS.index("esi_deduction"),
}

# Currency columns (for formatting)
CURRENCY_COLS = {
    "annual_ctc", "fixed_monthly", "basic", "hra", "ca", "sa", "gross",
    "per_day_salary", "lop", "pf_deduction", "pt_deduction", "esi_deduction", "net_pay",
}


# ══════════════════════════════════════════════════════════════
# SALARY CALCULATION HELPERS
# ══════════════════════════════════════════════════════════════

def _round2(val: Decimal) -> Decimal:
    """Round to 2 decimal places, half-up."""
    return val.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _parse_bool(val) -> bool:
    """
    Normalize boolean values read back from Excel.
    Excel stores booleans as 'Yes'/'No' strings (written by _build_export_workbook).
    Must handle True/False natives too for defensive safety.
    """
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.strip().lower() in ("yes", "true", "1")
    return bool(val)


def compute_pt(gross: Decimal) -> Decimal:
    """
    Professional Tax based on AP/Telangana monthly gross slab.
    PT_SLABS defined as (upper_bound, pt_amount); None upper_bound = no ceiling.
    """
    for upper, amount in PT_SLABS:
        if upper is None or gross <= upper:
            return amount
    return Decimal("0.00")  # unreachable but safe fallback


def compute_salary_components(annual_ctc: Decimal) -> dict:
    """
    Derive all salary components from annual CTC.
    Returns dict with Decimal values rounded to 2dp.
    """
    fixed   = _round2(annual_ctc / 12)
    basic   = _round2(fixed * Decimal("0.40"))
    hra     = _round2(basic * Decimal("0.50"))
    ca      = _round2(basic * Decimal("0.15"))
    sa      = _round2(fixed - basic - hra - ca)
    gross   = fixed  # always equals fixed by construction
    return {
        "fixed_monthly": fixed,
        "basic":         basic,
        "hra":           hra,
        "ca":            ca,
        "sa":            sa,
        "gross":         gross,
    }


def compute_deductions(
    gross: Decimal,
    basic: Decimal,
    calendar_days: int,
    absent_days: int,
    pf_enrolled: bool,
    esic_applicable: bool,
) -> dict:
    """
    Compute LOP, PF, PT, ESI deductions.

    Formula:
      lop     = gross / calendar_days * absent_days
      pf      = min(basic * 12%, 1800)   [if pf_enrolled]
      pt      = slab(gross)              [AP/Telangana Professional Tax]
      esi     = effective_gross * 0.75%  [if esic_applicable AND effective_gross <= 21000]
      net_pay = gross - lop - pf - pt - esi  (clamped to 0)
    """
    # LOP: proportional deduction for absent days
    if calendar_days > 0 and absent_days > 0:
        lop = _round2(gross / calendar_days * absent_days)
    else:
        lop = Decimal("0.00")

    # Effective earnings after LOP — used for ESI threshold check.
    # ESIC applies based on actual monthly wages earned, not gross CTC.
    effective_gross = gross - lop

    # PF: 12% of basic, capped at ₹1800
    if pf_enrolled:
        pf = _round2(min(basic * PF_RATE, PF_CAP))
    else:
        pf = Decimal("0.00")

    # PT: Professional Tax — AP/Telangana slab on gross salary
    pt = compute_pt(gross)

    # ESI: employee contribution 0.75% if effective_gross <= ₹21,000
    if esic_applicable and effective_gross <= ESI_THRESHOLD:
        esi = _round2(effective_gross * ESI_EMP_RATE)
    else:
        esi = Decimal("0.00")

    # Net pay — clamp to zero (avoids negative pay for high-LOP months)
    net_pay = _round2(max(Decimal("0.00"), gross - lop - pf - pt - esi))

    return {
        "lop":           lop,
        "pf_deduction":  pf,
        "pt_deduction":  pt,
        "esi_deduction": esi,
        "net_pay":       net_pay,
    }


# ══════════════════════════════════════════════════════════════
# DB QUERIES
# ══════════════════════════════════════════════════════════════

def _parse_weekly_off(weekly_off_str: Optional[str]) -> set[int]:
    """
    Parse employee's weekly_off string into weekday integers (Mon=0, Sun=6).
    Handles: "Sunday", "Saturday & Sunday", "Saturday, Sunday", etc.
    Defaults to {6} (Sunday only) if blank or unparseable.
    """
    WEEKDAY_MAP = {
        "monday": 0, "tuesday": 1, "wednesday": 2, "thursday": 3,
        "friday": 4, "saturday": 5, "sunday": 6,
    }
    if not weekly_off_str:
        return {6}
    days = set()
    for token in weekly_off_str.lower().replace("&", " ").replace(",", " ").split():
        if token in WEEKDAY_MAP:
            days.add(WEEKDAY_MAP[token])
    return days if days else {6}


async def _fetch_employees(
    db: asyncpg.Connection,
    year: int,
    month: int,
    branch_id: Optional[int],
    employee_ids: Optional[list[int]],
) -> list[dict]:
    """
    Pull payroll data for the month.

    Present day logic:
      present_days = DB present (punched in / approved leave / regularization)
                   + weekly-off days in month (Sundays etc — auto present)
                   + public holidays in month (auto present via daily_summary rows)
      absent_days  = eligible_days - present_days  (genuine LOP only)
      lop          = gross / eligible_days * absent_days

    Sandwich adjustment (when HR applies):
      Sandwich days are Sundays/holidays that sit between two leave periods.
      They are normally auto-present. When HR applies sandwich they become LOP.
      Fix: subtract sandwich days from working_days → lop_days increases → pay deducted.

    Pro-rata for new joiners: eligible period starts from date_of_joining.

    Performance: all per-employee data fetched in 3 batch queries (not N per employee).
    """
    cal_days    = calendar.monthrange(year, month)[1]
    month_start = date(year, month, 1)
    month_end   = date(year, month, cal_days)

    # ── Build employee WHERE clause ────────────────────────────
    filters = ["e.is_active = TRUE", "e.annual_ctc IS NOT NULL"]
    params  = []
    p = 1

    if branch_id:
        filters.append(f"e.branch_id = ${p}")
        params.append(branch_id)
        p += 1

    if employee_ids:
        filters.append(f"e.id = ANY(${p})")
        params.append(employee_ids)
        p += 1

    where = " AND ".join(filters)

    rows = await db.fetch(
        f"""
        SELECT
            COALESCE(e.emp_id, 'EMP-' || e.id::text)    AS emp_id,
            e.id                                        AS employee_id,
            u.full_name                                 AS employee_name,
            e.designation,
            e.department,
            b.name                                      AS branch,
            e.date_of_joining,
            e.weekly_off,
            e.bank_name,
            e.bank_account,
            e.bank_ifsc,
            e.pan_number,
            e.uan_number,
            e.pf_enrolled,
            e.esic_applicable,
            e.annual_ctc,
            e.user_id                                   AS user_id

        FROM employees e
        JOIN users u         ON u.id = e.user_id
        LEFT JOIN branches b ON b.id = e.branch_id

        WHERE {where}
        ORDER BY u.full_name
        """,
        *params,
    )

    if not rows:
        return []

    # ── Batch query 1: all daily_summary rows for all employees in month ──
    all_user_ids    = [r["user_id"]    for r in rows]
    all_employee_ids = [r["employee_id"] for r in rows]

    ds_all = await db.fetch(
        """
        SELECT
            ds.user_id,
            ds.work_date,
            ds.payroll_status,
            ds.status,
            lr.leave_type
        FROM daily_summary ds
        LEFT JOIN leave_requests lr ON lr.id = ds.leave_request_id
        WHERE ds.user_id  = ANY($1::int[])
          AND ds.work_date BETWEEN $2 AND $3
        """,
        all_user_ids, month_start, month_end,
    )
    # ds_by_user: user_id → {work_date → {payroll_status, status, leave_type}}
    ds_by_user: dict[int, dict] = {}
    for r in ds_all:
        ds_by_user.setdefault(r["user_id"], {})[r["work_date"]] = {
            "payroll_status": r["payroll_status"],
            "status":         r["status"],
            "leave_type":     r["leave_type"],
        }

    # ── Batch query 2: leave balances for all employees ────────
    bal_all = await db.fetch(
        """
        SELECT employee_id, total_paid_days, used_paid_days, remaining_paid_days
        FROM leave_balances
        WHERE employee_id = ANY($1::int[]) AND year = $2
        """,
        all_employee_ids, year,
    )
    bal_by_emp = {r["employee_id"]: r for r in bal_all}

    # ── Batch query 3: sandwich decisions for all employees ────
    payroll_month_date = date(year, month, 1)
    sandwich_all = await db.fetch(
        """
        SELECT employee_id, sandwich_days_detected, sandwich_applied
        FROM payroll_sandwich_reviews
        WHERE employee_id = ANY($1::int[]) AND payroll_month = $2
        """,
        all_employee_ids, payroll_month_date,
    )
    sandwich_by_emp = {
        r["employee_id"]: r["sandwich_days_detected"]
        for r in sandwich_all
        if r["sandwich_applied"]
    }

    today     = date.today()
    result    = []

    for r in rows:
        emp    = dict(r)
        annual = Decimal(str(emp["annual_ctc"]))
        sal    = compute_salary_components(annual)

        count_end      = min(month_end, today - timedelta(days=1))
        doj            = emp.get("date_of_joining")
        eligible_start = doj if doj and doj > month_start else month_start

        ds_map           = ds_by_user.get(emp["user_id"], {})
        weekly_off_days  = _parse_weekly_off(emp.get("weekly_off"))

        working_days = 0
        leaves_taken = 0   # approved paid leave days for display
        d = eligible_start
        while d <= count_end:
            if d in ds_map:
                if ds_map[d]["payroll_status"] == "present":
                    working_days += 1
                    if ds_map[d]["status"] == "leave" and ds_map[d]["leave_type"] == "paid":
                        leaves_taken += 1
            else:
                # No DB row — auto-present for weekly offs only.
                # Holidays already have DS rows from the holiday calendar feature.
                if d.weekday() in weekly_off_days:
                    working_days += 1
            d += timedelta(days=1)

        eligible_days = max(0, (count_end - eligible_start).days + 1) if count_end >= eligible_start else 0
        working_days  = min(working_days, eligible_days)

        # ── Sandwich adjustment ────────────────────────────────
        # Sandwich days are Sundays/holidays between two leave blocks.
        # Normally auto-present (no DS row, counted via weekly_off branch above).
        # When HR applies sandwich, each day is treated as:
        #   - Paid leave   → if employee has remaining paid balance
        #                    (working_days stays, balance consumed, leaves_taken++)
        #   - Unpaid/LOP   → if balance is exhausted
        #                    (working_days--, lop increases, salary deducted)
        leave_bal    = bal_by_emp.get(emp["employee_id"])
        paid_balance = int(leave_bal["remaining_paid_days"]) if leave_bal else 0

        sandwich_days = sandwich_by_emp.get(emp["employee_id"], 0)
        if sandwich_days > 0:
            sandwich_days   = min(sandwich_days, working_days)  # clamp to actual present days
            paid_sandwich   = min(sandwich_days, paid_balance)  # days covered by balance
            unpaid_sandwich = sandwich_days - paid_sandwich     # remaining → LOP

            leaves_taken  += paid_sandwich    # display: paid leave days used
            paid_balance  -= paid_sandwich    # track balance after sandwich
            working_days  -= unpaid_sandwich  # unpaid sandwich → absent → LOP

        lop_days = eligible_days - working_days

        available_leaves     = paid_balance   # reflects sandwich consumption if any
        carry_forward_leaves = available_leaves

        eff_cal = eligible_days if eligible_days < cal_days else cal_days
        per_day = _round2(sal["gross"] / eff_cal) if eff_cal > 0 else Decimal("0.00")

        ded = compute_deductions(
            gross           = sal["gross"],
            basic           = sal["basic"],
            calendar_days   = eff_cal,
            absent_days     = lop_days,
            pf_enrolled     = bool(emp["pf_enrolled"]),
            esic_applicable = bool(emp["esic_applicable"]),
        )

        emp.update(sal)
        emp.update(ded)
        emp["calendar_days"]        = eff_cal
        emp["present_days"]         = working_days
        emp["leaves_taken"]         = leaves_taken
        emp["lop_days"]             = lop_days
        emp["available_leaves"]     = available_leaves
        emp["carry_forward_leaves"] = carry_forward_leaves
        emp["per_day_salary"]       = per_day
        emp["annual_ctc"]           = annual
        result.append(emp)

    return result


# ══════════════════════════════════════════════════════════════
# EXCEL HELPERS
# ══════════════════════════════════════════════════════════════

def _style_header(ws, num_cols: int):
    """Apply header row styling."""
    header_fill   = PatternFill("solid", fgColor="041553")
    header_font   = Font(bold=True, color="FFFFFF", size=10)
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align

    ws.row_dimensions[1].height = 32


def _style_data_row(ws, row_idx: int, num_cols: int, editable_fill, locked_fill):
    """Apply alternating row colours; highlight editable cells."""
    alt_fill = PatternFill("solid", fgColor="F8F9FC") if row_idx % 2 == 0 else None

    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border    = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

        col_key = COLUMNS[col_idx - 1]
        if (col_idx - 1) in EDITABLE_COLS:
            cell.fill = editable_fill
        elif alt_fill:
            cell.fill = alt_fill


def _build_export_workbook(
    employees: list[dict],
    year: int,
    month: int,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Payroll {year}-{month:02d}"

    editable_fill = PatternFill("solid", fgColor="FFF9E6")  # light yellow = editable
    locked_fill   = PatternFill("solid", fgColor="F0F0F0")

    _style_header(ws, len(COLUMNS))

    for emp in employees:
        row_data = []
        for col in COLUMNS:
            val = emp.get(col)
            if isinstance(val, Decimal):
                val = float(val)
            elif isinstance(val, bool):
                val = "Yes" if val else "No"
            row_data.append(val)
        ws.append(row_data)

    # Style data rows
    for row_idx in range(2, len(employees) + 2):
        _style_data_row(ws, row_idx, len(COLUMNS), editable_fill, locked_fill)

    # Column widths
    col_widths = {
        "emp_id": 10, "employee_name": 22, "designation": 18, "department": 16,
        "branch": 14, "bank_name": 16, "bank_account": 18, "bank_ifsc": 14,
        "pan_number": 12, "pf_enrolled": 10, "esic_applicable": 12,
        "annual_ctc": 14, "fixed_monthly": 13, "basic": 10, "hra": 10,
        "ca": 10, "sa": 16, "gross": 12, "calendar_days": 12,
        "present_days": 11, "absent_days": 10, "lop": 13,
        "pf_deduction": 12, "pt_deduction": 12, "esi_deduction": 12, "net_pay": 13,
    }
    for idx, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = col_widths.get(col, 12)

    # Add legend note below data
    legend_row = len(employees) + 3
    ws.cell(row=legend_row, column=1,
            value="🟡 Yellow columns = HR editable | All others are calculated from DB")
    ws.cell(row=legend_row, column=1).font = Font(italic=True, color="888888", size=9)

    # Freeze header row
    ws.freeze_panes = "A2"

    return wb


# ══════════════════════════════════════════════════════════════
# PDF PAYSLIP HELPER
# ══════════════════════════════════════════════════════════════

def _generate_payslip_pdf(emp: dict, year: int, month: int) -> bytes:
    """
    Generate payslip PDF matching the Sri Dhanyadhathri reference layout:
      - Header  : Company name (large, blue) + address + phone/email, all in a bordered box
      - Title   : "Payslip for the month of Month,Year" centered, bordered
      - Info    : 4-col bordered table (label | value | label | value), plain black borders
      - Earnings: Bordered table with bold "Earnings:" / "Deductions:" headers
                  Rows: Fixed, Basic, HRA, CA, SA | PF, PT, ESI
                  Footer row: Gross Salary | Total Deductions (bold)
      - Net Pay : Single bold row spanning full width
      - Footer  : Auto-generated note, centered, small
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    BLACK  = colors.HexColor("#000000")
    BLUE   = colors.HexColor("#1a3a8f")   # company name blue
    WHITE  = colors.white
    LGREY  = colors.HexColor("#f5f5f5")

    THIN   = colors.HexColor("#000000")
    month_name = calendar.month_name[month]

    def fmt(val):
        """Format as Rs X,XXX.XX — uses Rs prefix (avoids rupee glyph encoding issues)."""
        if val is None:
            return ""
        try:
            return f"Rs {float(val):,.2f}"
        except Exception:
            return str(val)

    def fval(key):
        return float(emp.get(key) or 0)

    # ── Pull values ───────────────────────────────────────────
    gross        = fval("gross")
    fixed        = fval("fixed_monthly")
    basic        = fval("basic")
    hra          = fval("hra")
    ca           = fval("ca")
    sa           = fval("sa")
    lop          = fval("lop")
    pf           = fval("pf_deduction")
    pt           = fval("pt_deduction")
    esi          = fval("esi_deduction")
    net          = fval("net_pay")
    lop_days     = int(emp.get("lop_days")             or 0)
    leaves_taken = int(emp.get("leaves_taken")         or 0)
    present      = int(emp.get("present_days")         or 0)
    cal          = int(emp.get("calendar_days")        or 0)
    avail_leaves = int(emp.get("available_leaves")     or 0)
    carry_fwd    = int(emp.get("carry_forward_leaves") or 0)
    per_day      = fval("per_day_salary")
    total_ded    = lop + pf + pt + esi

    doj = emp.get("date_of_joining")
    if doj:
        try:
            doj_str = doj.strftime("%d-%m-%Y")
        except Exception:
            doj_str = str(doj)
    else:
        doj_str = "—"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=12*mm, bottomMargin=12*mm,
    )

    styles = getSampleStyleSheet()
    W = A4[0] - 30*mm  # usable width

    def s(name, **kw):
        return ParagraphStyle(name, parent=styles["Normal"], **kw)

    # Style shortcuts
    co_name   = s("co",   fontSize=16, textColor=BLUE,  fontName="Helvetica-Bold", leading=20)
    co_addr   = s("ca",   fontSize=7,  textColor=BLACK, leading=10)
    co_cont   = s("cc",   fontSize=7,  textColor=BLACK, leading=10)
    title_s   = s("ti",   fontSize=11, textColor=BLACK, fontName="Helvetica-Bold", alignment=TA_CENTER)
    lbl_s     = s("lb",   fontSize=9,  textColor=BLACK)
    val_s     = s("vl",   fontSize=9,  textColor=BLACK, fontName="Helvetica-Bold")
    hdr_s     = s("hd",   fontSize=10, textColor=BLACK, fontName="Helvetica-Bold")
    earn_lbl  = s("el",   fontSize=9,  textColor=BLACK)
    earn_val  = s("ev",   fontSize=9,  textColor=BLACK, alignment=TA_RIGHT)
    bold_lbl  = s("bl",   fontSize=9,  textColor=BLACK, fontName="Helvetica-Bold")
    bold_val  = s("bv",   fontSize=9,  textColor=BLACK, fontName="Helvetica-Bold", alignment=TA_RIGHT)
    net_lbl   = s("nl",   fontSize=10, textColor=BLACK, fontName="Helvetica-Bold")
    net_val   = s("nv",   fontSize=10, textColor=BLACK, fontName="Helvetica-Bold", alignment=TA_RIGHT)
    footer_s  = s("ft",   fontSize=7,  textColor=colors.HexColor("#666666"), alignment=TA_CENTER)

    BOX   = ("BOX",      (0,0), (-1,-1), 0.8, THIN)
    GRID  = ("INNERGRID",(0,0), (-1,-1), 0.5, THIN)
    VPAD  = [("TOPPADDING",(0,0),(-1,-1),4), ("BOTTOMPADDING",(0,0),(-1,-1),4)]
    HPAD  = [("LEFTPADDING",(0,0),(-1,-1),6), ("RIGHTPADDING",(0,0),(-1,-1),6)]
    VMID  = ("VALIGN",(0,0),(-1,-1),"MIDDLE")

    story = []

    # ══════════════════════════════════════════════════════════
    # 1. HEADER — Logo left, Company name + address right, blue border
    # ══════════════════════════════════════════════════════════
    from reportlab.platypus import Image as RLImage

    from pathlib import Path as _Path
    LOGO_PATH = str(_Path(__file__).resolve().parent.parent / "frontend" / "icon-51.png")
    try:
        logo = RLImage(LOGO_PATH, width=18*mm, height=18*mm)
    except Exception:
        logo = Paragraph("", s("empty"))

    addr_block = Table(
        [
            [Paragraph("GodavariKrishna Group", co_name)],
            [Paragraph("#24-28/1-7/1, Vijayawada Municipal Corporation, Revenue Ward 43, NTR District, Andhra Pradesh - 520003.", co_addr)],
            [Paragraph(
                '<font size="8">&#9990;</font>  +91-8121033925'
                '                                        '
                '<font size="8">&#9993;</font>  admin@gkgroup.com',
                co_cont
            )],
        ],
        colWidths=[W - 22*mm],
    )
    addr_block.setStyle(TableStyle([
        ("LEFTPADDING",  (0,0),(-1,-1), 0),
        ("RIGHTPADDING", (0,0),(-1,-1), 0),
        ("TOPPADDING",   (0,0),(-1,-1), 1),
        ("BOTTOMPADDING",(0,0),(-1,-1), 1),
    ]))

    hdr_tbl = Table(
        [[logo, addr_block]],
        colWidths=[22*mm, W - 22*mm],
    )
    BLUE_BORDER = colors.HexColor("#1a3a8f")
    hdr_tbl.setStyle(TableStyle([
        ("BOX",          (0,0),(-1,-1), 1.5, BLUE_BORDER),
        ("VALIGN",       (0,0),(-1,-1), "MIDDLE"),
        ("LEFTPADDING",  (0,0),(0,0),   4),
        ("RIGHTPADDING", (0,0),(0,0),   4),
        ("LEFTPADDING",  (1,0),(1,0),   6),
        ("TOPPADDING",   (0,0),(-1,-1), 6),
        ("BOTTOMPADDING",(0,0),(-1,-1), 6),
    ]))
    story.append(hdr_tbl)

    # ══════════════════════════════════════════════════════════
    # 2. TITLE ROW
    # ══════════════════════════════════════════════════════════
    title_tbl = Table(
        [[Paragraph(f"Payslip for the month of {month_name},{year}", title_s)]],
        colWidths=[W],
    )
    title_tbl.setStyle(TableStyle([
        BOX, VMID,
        ("TOPPADDING",(0,0),(-1,-1),6), ("BOTTOMPADDING",(0,0),(-1,-1),6),
    ]))
    story.append(title_tbl)

    # ══════════════════════════════════════════════════════════
    # 3. EMPLOYEE INFO GRID
    # 4-column: label | value | label | value
    # ══════════════════════════════════════════════════════════
    def irow(l1, v1, l2="", v2=""):
        return [
            Paragraph(l1 + (":" if l1 else ""), lbl_s),
            Paragraph(str(v1) if v1 else "—", val_s),
            Paragraph(l2 + (":" if l2 else ""), lbl_s),
            Paragraph(str(v2) if v2 else "", val_s),
        ]

    cw = W / 4
    info_data = [
        irow("Emp Name",      emp.get("employee_name") or "—", "UAN",          emp.get("uan_number") or "—"),
        irow("DOJ",           doj_str,                          "Bank Name",    emp.get("bank_name") or "—"),
        irow("Designation",   emp.get("designation") or "—",   "Bank A/C no",  emp.get("bank_account") or "—"),
        irow("Department",    emp.get("department") or "—",     "LOP",          str(lop_days)),
        irow("No of days worked", str(present),                 "",             ""),
    ]
    info_tbl = Table(info_data, colWidths=[cw * 0.6, cw * 0.9, cw * 0.6, cw * 0.9])
    info_tbl.setStyle(TableStyle([
        BOX, GRID, VMID,
        *VPAD, *HPAD,
    ]))
    story.append(info_tbl)

    # ══════════════════════════════════════════════════════════
    # 4. EARNINGS / DEDUCTIONS TABLE
    # ══════════════════════════════════════════════════════════
    cw2 = W / 4

    def erow(el, ev, dl="", dv=""):
        return [
            Paragraph(el, earn_lbl),
            Paragraph(fmt(ev) if ev != "" else "", earn_val),
            Paragraph(dl + (" :" if dl else ""), earn_lbl),
            Paragraph(fmt(dv) if dv != "" else "", earn_val),
        ]

    sal_data = [
        # Header row
        [
            Paragraph("Earnings :", hdr_s), "",
            Paragraph("Deductions:", hdr_s), "",
        ],
        erow("Fixed",               fixed,  "PF",  pf),
        erow("Basic",               basic,  "PT",  pt),
        erow("HRA",                 hra,    "ESI", esi),
        erow("Conveyance Allowance",ca,     "",    ""),
        erow("Special Allowance",   sa,     "",    ""),
        # Gross / Total Deductions row
        [
            Paragraph("Gross Salary", bold_lbl),
            Paragraph(fmt(gross),     bold_val),
            Paragraph("Total Deductions:", bold_lbl),
            Paragraph(fmt(total_ded), bold_val),
        ],
    ]

    sal_tbl = Table(sal_data, colWidths=[cw2 * 1.15, cw2 * 0.85, cw2 * 1.15, cw2 * 0.85])
    sal_tbl.setStyle(TableStyle([
        BOX, GRID, VMID,
        *VPAD, *HPAD,
        # Header row: span label across 2 cols each side
        ("SPAN",         (0,0),(1,0)),
        ("SPAN",         (2,0),(3,0)),
        ("BACKGROUND",   (0,0),(-1,0), LGREY),
        ("FONTNAME",     (0,0),(-1,0), "Helvetica-Bold"),
        # Gross row bold
        ("FONTNAME",     (0,-1),(-1,-1), "Helvetica-Bold"),
        ("BACKGROUND",   (0,-1),(-1,-1), LGREY),
    ]))
    story.append(sal_tbl)

    # ══════════════════════════════════════════════════════════
    # 5. NET PAY ROW
    # ══════════════════════════════════════════════════════════
    net_tbl = Table(
        [[Paragraph("NET PAY", net_lbl), Paragraph(fmt(net), net_val)]],
        colWidths=[W * 0.5, W * 0.5],
    )
    net_tbl.setStyle(TableStyle([
        BOX, VMID,
        ("TOPPADDING",(0,0),(-1,-1),6), ("BOTTOMPADDING",(0,0),(-1,-1),6),
        ("LEFTPADDING",(0,0),(-1,-1),8), ("RIGHTPADDING",(0,0),(-1,-1),8),
    ]))
    story.append(net_tbl)
    story.append(Spacer(1, 6*mm))

    # ══════════════════════════════════════════════════════════
    # 6. FOOTER
    # ══════════════════════════════════════════════════════════
    story.append(Paragraph(
        "This is a computer-generated payslip and does not require a signature.  |  Generated by GodavariKrishna HRMS",
        footer_s,
    ))

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════════
# ENDPOINTS
# ══════════════════════════════════════════════════════════════

@router.get("/export")
async def export_payroll(
    year:         int            = Query(..., ge=2020, le=2100),
    month:        int            = Query(..., ge=1,    le=12),
    branch_id:    Optional[int]  = Query(None),
    employee_ids: Optional[str]  = Query(None, description="Comma-separated employee IDs"),
    _hr: dict    = Depends(require_hr),
    db: asyncpg.Connection = Depends(get_db),
):
    """
    Export pre-filled payroll Excel for the given month.
    HR cross-checks, corrects if needed, then uploads back via /process.

    Query params:
      year, month      — payroll period
      branch_id        — filter by branch (optional)
      employee_ids     — comma-separated employee IDs (optional, overrides branch_id)
    """
    # Parse employee_ids
    emp_id_list = None
    if employee_ids:
        try:
            emp_id_list = [int(x.strip()) for x in employee_ids.split(",") if x.strip()]
        except ValueError:
            raise HTTPException(400, "employee_ids must be comma-separated integers")

    employees = await _fetch_employees(db, year, month, branch_id, emp_id_list)

    if not employees:
        raise HTTPException(404, "No employees found for the given filters")

    wb = _build_export_workbook(employees, year, month)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"payroll_{year}_{month:02d}.xlsx"
    logger.info(
        "Payroll export: year=%d month=%d employees=%d by hr=%s",
        year, month, len(employees), _hr.get("id"),
    )

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/process")
async def process_payroll(
    year:  int = Query(..., ge=2020, le=2100),
    month: int = Query(..., ge=1,    le=12),
    file:  UploadFile = File(...),
    _hr: dict = Depends(require_hr),
    db: asyncpg.Connection = Depends(get_db),
):
    """
    Accept the corrected payroll Excel, recalculate using HR's values, and return:
      - A ZIP file containing:
          payroll_summary_YYYY_MM.xlsx  — full summary sheet
          payslips/EMP-XXXX.pdf        — individual PDF payslip per employee

    Excel wins over DB — whatever HR has put in the sheet is used as-is.
    Recalculates: net_pay = gross - lop - pf_deduction - esi_deduction
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files accepted")

    content = await file.read()
    try:
        wb_in = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Invalid Excel file: {e}")

    ws = wb_in.active
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        raise HTTPException(400, "Excel has no data rows")

    # Validate header matches expected COLUMNS
    actual_headers = [str(h).strip() if h else "" for h in rows[0]]
    expected_headers = HEADERS
    if actual_headers != expected_headers:
        raise HTTPException(
            400,
            f"Excel header mismatch. Expected columns: {expected_headers}. "
            f"Got: {actual_headers}"
        )

    processed = []
    errors    = []

    for row_num, row in enumerate(rows[1:], start=2):
        # Skip rows with no employee name (empty/legend rows)
        # emp_id can be NULL in DB so don't use it as the skip check
        if not row[1]:  # employee_name is col index 1
            continue

        emp = dict(zip(COLUMNS, row))

        # Normalize booleans — Excel stores them as "Yes"/"No" strings
        emp["pf_enrolled"]     = _parse_bool(emp.get("pf_enrolled"))
        emp["esic_applicable"] = _parse_bool(emp.get("esic_applicable"))

        # Basic validation
        try:
            gross        = Decimal(str(emp.get("gross") or 0))
            lop          = Decimal(str(emp.get("lop")   or 0))
            pf           = Decimal(str(emp.get("pf_deduction")  or 0))
            pt           = Decimal(str(emp.get("pt_deduction")  or 0))
            esi          = Decimal(str(emp.get("esi_deduction") or 0))
            present_days = int(emp.get("present_days") or 0)
            lop_days     = int(emp.get("lop_days")     or 0)
            leaves_taken = int(emp.get("leaves_taken") or 0)
            cal_days     = int(emp.get("calendar_days") or calendar.monthrange(year, month)[1])
        except (ValueError, TypeError) as e:
            errors.append(f"Row {row_num} ({emp.get('emp_id')}): invalid number — {e}")
            continue

        # Sanity checks
        if present_days + lop_days > cal_days:
            errors.append(
                f"Row {row_num} ({emp.get('emp_id')}): "
                f"present ({present_days}) + lop_days ({lop_days}) > calendar days ({cal_days})"
            )
            continue

        if lop < 0 or pf < 0 or pt < 0 or esi < 0:
            errors.append(f"Row {row_num} ({emp.get('emp_id')}): deductions cannot be negative")
            continue

        # Recalculate net pay from HR's numbers (Excel is truth)
        # Clamp to zero — prevents negative net pay when HR enters large LOP
        net_pay = _round2(max(Decimal("0.00"), gross - lop - pf - pt - esi))

        emp["gross"]          = gross
        emp["lop"]            = lop
        emp["pf_deduction"]   = pf
        emp["pt_deduction"]   = pt
        emp["esi_deduction"]  = esi
        emp["net_pay"]        = net_pay
        emp["present_days"]   = present_days
        emp["lop_days"]       = lop_days
        emp["leaves_taken"]   = leaves_taken
        emp["calendar_days"]  = cal_days

        processed.append(emp)

    if errors:
        raise HTTPException(422, {"message": "Validation errors in Excel", "errors": errors})

    if not processed:
        raise HTTPException(400, "No valid employee rows found in Excel")

    # ── Build output ZIP ──────────────────────────────────────
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:

        # 1. Summary Excel
        summary_wb = _build_summary_workbook(processed, year, month)
        summary_buf = io.BytesIO()
        summary_wb.save(summary_buf)
        zf.writestr(f"payroll_summary_{year}_{month:02d}.xlsx", summary_buf.getvalue())

        # 2. Individual payslips (HTML → print as PDF)
        for emp in processed:
            payslip_bytes = _generate_payslip_pdf(emp, year, month)
            safe_name = str(emp.get("emp_id") or emp.get("employee_name") or "unknown").replace("/", "-")
            zf.writestr(f"payslips/{safe_name}.pdf", payslip_bytes)

    zip_buf.seek(0)
    filename = f"payroll_{year}_{month:02d}_processed.zip"

    logger.info(
        "Payroll processed: year=%d month=%d employees=%d by hr=%s",
        year, month, len(processed), _hr.get("id"),
    )

    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_summary_workbook(employees: list[dict], year: int, month: int) -> Workbook:
    """Build the final payroll summary Excel (same format as export)."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Summary {year}-{month:02d}"

    editable_fill = PatternFill("solid", fgColor="FFFFFF")
    locked_fill   = PatternFill("solid", fgColor="F8F9FC")

    _style_header(ws, len(COLUMNS))

    for emp in employees:
        row_data = []
        for col in COLUMNS:
            val = emp.get(col)
            if isinstance(val, Decimal):
                val = float(val)
            elif isinstance(val, bool):
                val = "Yes" if val else "No"
            row_data.append(val)
        ws.append(row_data)

    for row_idx in range(2, len(employees) + 2):
        _style_data_row(ws, row_idx, len(COLUMNS), editable_fill, locked_fill)

    # Totals row
    total_row = len(employees) + 2
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).font = Font(bold=True)

    currency_col_indices = [i + 1 for i, col in enumerate(COLUMNS) if col in CURRENCY_COLS]
    for col_idx in currency_col_indices:
        col_letter = get_column_letter(col_idx)
        ws.cell(
            row=total_row, column=col_idx,
            value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})",
        ).font = Font(bold=True)

    ws.freeze_panes = "A2"
    return wb